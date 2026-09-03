from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import subprocess
import time
import zipfile
from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import load_workbook

from app.domain.cleaner_registry import CleanerRelease
from app.domain.quick_analysis import QuickAnalysisArtifact
from app.infrastructure.child_process_environment import isolated_child_environment
from app.infrastructure.existing_cleaner_runner import ExistingCleanerRunner

ProcessRunner = Callable[..., subprocess.CompletedProcess[str]]


@dataclass(frozen=True, slots=True)
class QuickPatRunResult:
    parameter_count: int
    record_count: int | None
    summary: dict[str, Any]
    artifacts: tuple[QuickAnalysisArtifact, ...]
    stdout_tail: str


class QuickPatRunner:
    """Invoke released FT raw PAT or CP Cleaner + package-owned PAT."""

    def __init__(
        self,
        process_runner: ProcessRunner = subprocess.run,
        cleaner_runner: ExistingCleanerRunner | None = None,
    ) -> None:
        self._process_runner = process_runner
        self._cleaner_runner = cleaner_runner or ExistingCleanerRunner(
            process_runner=process_runner
        )

    def run_release(
        self,
        *,
        release: CleanerRelease,
        input_directory: str | Path,
        output_root: str | Path,
        source_manifest_json: str,
        source_manifest_sha256: str,
    ) -> QuickPatRunResult:
        if release.test_stage == "CP":
            return self._run_cp_release(
                release=release,
                input_directory=input_directory,
                output_root=output_root,
                source_manifest_json=source_manifest_json,
                source_manifest_sha256=source_manifest_sha256,
            )
        approved_ft = {
            "JIEQUN_FT_QUICK_PAT_PYZ": (
                "JIEQUN",
                "JIEQUN_UNIFIED_CSV_DIRECTORY_V1",
            ),
            "RIYUEXIN_FT_QUICK_PAT_PYZ": (
                "RIYUEXIN",
                "RIYUEXIN_RAW_XLSX_DIRECTORY_V1",
            ),
            "RIYUEGUANG_FT_QUICK_PAT_PYZ": (
                "RIYUEGUANG",
                "RIYUEGUANG_RAW_XLSX_DIRECTORY_V1",
            ),
            "DIANJI_FT_QUICK_PAT_PYZ": (
                "DIANJI",
                "DIANJI_REGISTERED_RAW_DIRECTORY_V1",
            ),
            "JIJIA_FT_QUICK_PAT_PYZ": (
                "JIJIA",
                "JIJIA_STS8203_CSV_DIRECTORY_V1",
            ),
        }
        expected_ft = approved_ft.get(release.adapter_code)
        if (
            release.test_stage != "FT"
            or expected_ft is None
            or (release.factory_code, release.input_contract_version) != expected_ft
            or release.output_contract_version != "FT_PAT_RESULT_V1"
        ):
            raise ValueError("released tool is not an approved FT Quick PAT adapter")
        source = Path(input_directory).resolve()
        package = Path(release.artifact_uri).resolve()
        runtime = Path(release.runtime_uri).resolve()
        target = Path(output_root).resolve()
        if not source.exists() or not (source.is_dir() or source.is_file()):
            raise FileNotFoundError(f"Quick PAT input source is unavailable: {source}")
        for required in (runtime, package):
            if not required.is_file():
                raise FileNotFoundError(f"Quick PAT runtime is unavailable: {required}")
        if _file_sha256(package) != release.code_checksum.lower():
            raise RuntimeError(
                f"PAT package checksum differs from released contract: {package}"
            )
        actual_manifest_sha = hashlib.sha256(
            source_manifest_json.encode("utf-8")
        ).hexdigest()
        if actual_manifest_sha != source_manifest_sha256.lower():
            raise RuntimeError("source Manifest JSON does not match its SHA-256")
        manifest = json.loads(source_manifest_json)
        target.mkdir(parents=True, exist_ok=True)
        started = time.perf_counter()
        engine_source = source
        staged_source: Path | None = None
        archive_input = source.is_file() and source.suffix.lower() in {".zip", ".7z"}
        try:
            if source.is_file():
                staged_source = target / ".single-source"
                staged_source.mkdir()
                if archive_input:
                    _extract_ft_archive(
                        source,
                        staged_source,
                        allowed_suffixes=_ft_raw_suffixes(release.adapter_code),
                        max_total_bytes=_quick_archive_max_extracted_bytes(),
                    )
                else:
                    staged_file = staged_source / source.name
                    try:
                        staged_file.hardlink_to(source)
                    except OSError:
                        shutil.copy2(source, staged_file)
                engine_source = staged_source

            env = isolated_child_environment(
                {
                    "TMS_FT_PAT_PACKAGE": str(package),
                    "TMS_FT_PAT_INPUT": str(engine_source),
                    "TMS_FT_PAT_OUTPUT": str(target),
                    "TMS_FT_PAT_ADAPTER": release.adapter_code,
                    "PYTHONUTF8": "1",
                    "PYTHONIOENCODING": "utf-8",
                }
            )
            completed = self._process_runner(
                [str(runtime), "-c", _FT_QUICK_PAT_SCRIPT],
                cwd=str(package.parent),
                env=env,
                text=True,
                encoding="utf-8",
                errors="replace",
                capture_output=True,
                check=False,
                timeout=release.timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError(
                f"Quick PAT timed out after {release.timeout_seconds}s"
            ) from exc
        finally:
            if staged_source is not None and staged_source.exists():
                shutil.rmtree(staged_source)
        elapsed_seconds = time.perf_counter() - started
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "unknown PAT error")[-4000:]
            raise RuntimeError(
                f"released {release.factory_code} Quick PAT failed: {detail}"
            )

        reports = sorted(
            target.rglob("PAT_*.xlsx"), key=lambda item: str(item).casefold()
        )
        if len(reports) != 1:
            raise RuntimeError(
                f"Quick PAT output contract requires one PAT Excel, found {len(reports)}"
            )
        report = reports[0].resolve()
        parameter_rows = _read_pat_rows(report)
        counts = [
            int(row["count"])
            for row in parameter_rows
            if isinstance(row.get("count"), (int, float))
        ]
        engine_file_count, parsed_record_count, engine_parameter_count = (
            _parse_engine_summary(completed.stdout or "")
        )
        if not archive_input and engine_file_count != manifest.get("file_count"):
            raise RuntimeError(
                "PAT engine source count differs from the submitted Manifest: "
                f"{engine_file_count}!={manifest.get('file_count')}"
            )
        if engine_parameter_count != len(parameter_rows):
            raise RuntimeError(
                "PAT engine parameter count differs from the result workbook: "
                f"{engine_parameter_count}!={len(parameter_rows)}"
            )
        summary: dict[str, Any] = {
            "schema_version": 1,
            "analysis_type": "QUICK_PAT",
            "test_stage": "FT",
            "factory_code": release.factory_code,
            "input_contract": release.input_contract_version,
            "output_contract": release.output_contract_version,
            "formula_contract": "SIGMA_IQR_1_35_MEDIAN_PLUS_MINUS_6SIGMA_V1",
            "manifest_mode": manifest.get("mode"),
            "source_file_count": manifest.get("file_count"),
            "source_total_bytes": manifest.get("total_bytes"),
            "source_manifest_sha256": source_manifest_sha256,
            "engine_source_file_count": engine_file_count,
            "archive_input": archive_input,
            "parameter_count": len(parameter_rows),
            "record_count": parsed_record_count,
            "record_count_basis": "PAT_ENGINE_PARSED_ROWS",
            "maximum_parameter_value_count": max(counts) if counts else None,
            "minimum_parameter_value_count": min(counts) if counts else None,
            "elapsed_seconds": round(elapsed_seconds, 3),
            "parameters": parameter_rows,
        }
        manifest_path = target / "source_manifest.json"
        manifest_path.write_text(source_manifest_json, encoding="utf-8")
        summary_path = target / "pat_summary.json"
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        artifacts = tuple(
            _artifact(role, path)
            for role, path in (
                ("pat_report", report),
                ("pat_summary", summary_path),
                ("source_manifest", manifest_path),
            )
        )
        output_bytes = sum(
            path.stat().st_size for path in target.rglob("*") if path.is_file()
        )
        if output_bytes > release.max_output_bytes:
            raise RuntimeError(
                "Quick PAT output exceeds released limit: "
                f"{output_bytes}>{release.max_output_bytes}"
            )
        return QuickPatRunResult(
            len(parameter_rows),
            parsed_record_count,
            summary,
            artifacts,
            (completed.stdout or "")[-4000:],
        )

    def _run_cp_release(
        self,
        *,
        release: CleanerRelease,
        input_directory: str | Path,
        output_root: str | Path,
        source_manifest_json: str,
        source_manifest_sha256: str,
    ) -> QuickPatRunResult:
        approved = {
            "HUAHONG_CP_PYZ": ("HUAHONG", "CP_ARCHIVE_OR_TXT_V1", "CP_CSV_TRIPLET_V1"),
            "JETECH_CP_PYZ": ("JETECH", "CP_EXCEL_OR_ZIP_V1", "CP_STANDARD_CSV_TRIPLET_V1"),
            "LION_CP_PYZ": ("LION", "CP_EXCEL_OR_ZIP_V1", "CP_STANDARD_CSV_TRIPLET_V1"),
            "GUOYU_CP_PYZ": ("GUOYU", "CP_EXCEL_OR_ZIP_V1", "CP_STANDARD_CSV_TRIPLET_V1"),
        }
        expected = approved.get(release.adapter_code)
        if expected is None or (
            release.factory_code,
            release.input_contract_version,
            release.output_contract_version,
        ) != expected:
            raise ValueError("released tool is not an approved CP Quick PAT adapter")

        source = Path(input_directory).resolve()
        target = Path(output_root).resolve()
        package = Path(release.artifact_uri).resolve()
        runtime = Path(release.runtime_uri).resolve()
        if not source.exists() or not (source.is_dir() or source.is_file()):
            raise FileNotFoundError(f"Quick PAT input source is unavailable: {source}")
        for required in (runtime, package):
            if not required.is_file():
                raise FileNotFoundError(f"CP Quick PAT runtime is unavailable: {required}")
        if _file_sha256(package) != release.code_checksum.lower():
            raise RuntimeError(
                f"CP PAT package checksum differs from released contract: {package}"
            )
        actual_manifest_sha = hashlib.sha256(
            source_manifest_json.encode("utf-8")
        ).hexdigest()
        if actual_manifest_sha != source_manifest_sha256.lower():
            raise RuntimeError("source Manifest JSON does not match its SHA-256")
        manifest = json.loads(source_manifest_json)
        target.mkdir(parents=True, exist_ok=True)
        intermediate = (target / "cp_cleaner_intermediate").resolve()
        if intermediate.parent != target:
            raise RuntimeError("CP Quick PAT intermediate path escaped its job workspace")

        started = time.perf_counter()
        cleaner_result = self._cleaner_runner.run_release(
            release=release,
            inputs=(source,),
            output_root=intermediate,
        )
        cleaned = [item.path for item in cleaner_result.artifacts if item.role == "cleaned"]
        specs = [item.path for item in cleaner_result.artifacts if item.role == "spec"]
        env = isolated_child_environment(
            {
                "TMS_CP_PAT_PACKAGE": str(package),
                "TMS_CP_PAT_CLEANED_JSON": json.dumps(cleaned, ensure_ascii=False),
                "TMS_CP_PAT_SPEC_JSON": json.dumps(specs, ensure_ascii=False),
                "TMS_CP_PAT_OUTPUT": str(target),
                "PYTHONUTF8": "1",
                "PYTHONIOENCODING": "utf-8",
            }
        )
        try:
            completed = self._process_runner(
                [str(runtime), "-c", _CP_QUICK_PAT_SCRIPT],
                cwd=str(package.parent),
                env=env,
                text=True,
                encoding="utf-8",
                errors="replace",
                capture_output=True,
                check=False,
                timeout=release.timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError(
                f"CP Quick PAT timed out after {release.timeout_seconds}s"
            ) from exc
        finally:
            if intermediate.exists() and intermediate.parent == target:
                shutil.rmtree(intermediate)
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "unknown CP PAT error")[-4000:]
            raise RuntimeError(f"released CP Quick PAT failed: {detail}")

        reports = sorted(target.glob("PAT_CP_*.xlsx"), key=lambda item: item.name.casefold())
        if len(reports) != 1:
            raise RuntimeError(
                f"CP Quick PAT output contract requires one PAT Excel, found {len(reports)}"
            )
        engine = _parse_cp_engine_summary(completed.stdout or "")
        report = reports[0].resolve()
        parameter_rows = _read_pat_rows(report)
        if engine["parameter_count"] != len(parameter_rows):
            raise RuntimeError("CP PAT engine parameter count differs from result workbook")
        elapsed_seconds = time.perf_counter() - started
        summary: dict[str, Any] = {
            "schema_version": 1,
            "analysis_type": "QUICK_PAT",
            "test_stage": "CP",
            "factory_code": release.factory_code,
            "input_contract": release.input_contract_version,
            "output_contract": "CP_PAT_RESULT_V1",
            "formula_contract": engine["formula_contract"],
            "formula_provenance": "VDMOS_Tool_v5.6.html existing CP PAT",
            "manifest_mode": manifest.get("mode"),
            "source_file_count": manifest.get("file_count"),
            "source_total_bytes": manifest.get("total_bytes"),
            "source_manifest_sha256": source_manifest_sha256,
            "cleaned_file_count": engine["source_files"],
            "parameter_count": len(parameter_rows),
            "record_count": engine["record_count"],
            "record_count_basis": "CP_CLEANER_STANDARD_CSV_ROWS",
            "elapsed_seconds": round(elapsed_seconds, 3),
            "parameters": parameter_rows,
        }
        manifest_path = target / "source_manifest.json"
        manifest_path.write_text(source_manifest_json, encoding="utf-8")
        summary_path = target / "pat_summary.json"
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        artifacts = tuple(
            _artifact(role, path)
            for role, path in (
                ("pat_report", report),
                ("pat_summary", summary_path),
                ("source_manifest", manifest_path),
            )
        )
        output_bytes = sum(path.stat().st_size for path in target.rglob("*") if path.is_file())
        if output_bytes > release.max_output_bytes:
            raise RuntimeError(
                f"CP Quick PAT output exceeds released limit: {output_bytes}>{release.max_output_bytes}"
            )
        return QuickPatRunResult(
            len(parameter_rows),
            int(engine["record_count"]),
            summary,
            artifacts,
            (cleaner_result.stdout_tail + "\n" + (completed.stdout or ""))[-4000:],
        )


def _read_pat_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows: list[dict[str, Any]] = []
        keys = (
            "parameter",
            "count",
            "mean",
            "stddev",
            "minimum",
            "q1",
            "median",
            "q3",
            "maximum",
            "sigma",
            "lcl_calculated",
            "ucl_calculated",
            "lcl_before",
            "ucl_before",
            "lcl_after",
            "ucl_after",
            "updated",
        )
        for index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index <= 2:
                continue
            normalized = list(values[: len(keys)])
            normalized.extend([None] * (len(keys) - len(normalized)))
            parameter = str(normalized[0]).strip() if normalized[0] is not None else ""
            if not parameter:
                continue
            if not isinstance(normalized[1], (int, float)):
                raise TypeError(f"PAT Excel has an invalid count for {parameter}")
            rows.append(dict(zip(keys, normalized, strict=True)))
    finally:
        workbook.close()
    if not rows:
        raise RuntimeError("PAT Excel contains no parameter rows")
    return rows


def _parse_engine_summary(stdout: str) -> tuple[int, int, int]:
    matches = re.findall(
        r"PAT [^\r\n]*?:\s*文件=([\d,]+),\s*解析行=([\d,]+),\s*参数=(\d+)",
        stdout,
    )
    if len(matches) != 1:
        raise RuntimeError("PAT engine did not emit one auditable source summary")
    files, rows, parameters = matches[0]
    return int(files.replace(",", "")), int(rows.replace(",", "")), int(parameters)


def _parse_cp_engine_summary(stdout: str) -> dict[str, Any]:
    marked = [
        line.removeprefix("TMS_CP_PAT_SUMMARY=")
        for line in stdout.splitlines()
        if line.startswith("TMS_CP_PAT_SUMMARY=")
    ]
    if len(marked) != 1:
        raise RuntimeError("CP PAT engine did not emit one auditable summary")
    payload = json.loads(marked[0])
    required = {"source_files", "record_count", "parameter_count", "formula_contract", "report"}
    if not isinstance(payload, dict) or set(payload) != required:
        raise RuntimeError("CP PAT engine summary has an invalid schema")
    return payload


def _artifact(role: str, path: Path) -> QuickAnalysisArtifact:
    resolved = path.resolve()
    return QuickAnalysisArtifact(
        role,
        str(resolved),
        resolved.stat().st_size,
        _file_sha256(resolved),
    )


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _ft_raw_suffixes(adapter_code: str) -> tuple[str, ...]:
    mapping = {
        "JIEQUN_FT_QUICK_PAT_PYZ": (".csv",),
        "RIYUEXIN_FT_QUICK_PAT_PYZ": (".xlsx",),
        "RIYUEGUANG_FT_QUICK_PAT_PYZ": (".xlsx",),
        "DIANJI_FT_QUICK_PAT_PYZ": (".xls", ".xlsx", ".csv"),
        "JIJIA_FT_QUICK_PAT_PYZ": (".csv",),
    }
    try:
        return mapping[adapter_code]
    except KeyError as exc:
        raise ValueError(f"unsupported FT Quick PAT adapter: {adapter_code}") from exc


def _quick_archive_max_extracted_bytes() -> int:
    value = int(
        os.getenv("TMS_QUICK_ARCHIVE_MAX_EXTRACTED_BYTES", str(100 * 1024**3))
    )
    if value < 1:
        raise ValueError("TMS_QUICK_ARCHIVE_MAX_EXTRACTED_BYTES必须大于0")
    return value


def _archive_destination(root: Path, member_name: str) -> Path:
    normalized = member_name.replace("\\", "/")
    relative = PurePosixPath(normalized)
    if (
        not normalized
        or normalized.startswith("/")
        or relative.is_absolute()
        or any(part in {"", ".", ".."} for part in relative.parts)
        or (relative.parts and ":" in relative.parts[0])
    ):
        raise ValueError(f"压缩包包含无效路径: {member_name}")
    destination = (root / Path(*relative.parts)).resolve()
    resolved_root = root.resolve()
    if destination == resolved_root or resolved_root not in destination.parents:
        raise ValueError(f"压缩包文件超出临时目录: {member_name}")
    return destination


def _extract_ft_archive(
    source: Path,
    target: Path,
    *,
    allowed_suffixes: tuple[str, ...],
    max_total_bytes: int,
) -> tuple[Path, ...]:
    suffixes = {item.lower() for item in allowed_suffixes}
    extracted: list[Path] = []
    total_bytes = 0
    if source.suffix.lower() == ".zip":
        try:
            with zipfile.ZipFile(source) as archive:
                selected = [
                    item
                    for item in archive.infolist()
                    if not item.is_dir()
                    and PurePosixPath(item.filename.replace("\\", "/")).suffix.lower()
                    in suffixes
                ]
                if any(item.flag_bits & 0x1 for item in selected):
                    raise ValueError("FT压缩包不能带密码")
                total_bytes = sum(item.file_size for item in selected)
                if total_bytes > max_total_bytes:
                    raise ValueError("FT压缩包解压后超过快速分析解压容量上限")
                for item in selected:
                    destination = _archive_destination(target, item.filename)
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    with archive.open(item) as source_stream, destination.open("xb") as output:
                        shutil.copyfileobj(source_stream, output, length=1024 * 1024)
                    extracted.append(destination)
        except zipfile.BadZipFile as exc:
            raise ValueError("FT ZIP压缩包损坏或无法读取") from exc
    elif source.suffix.lower() == ".7z":
        import py7zr

        try:
            with py7zr.SevenZipFile(source, mode="r") as archive:
                if archive.needs_password():
                    raise ValueError("FT压缩包不能带密码")
                selected = [
                    item
                    for item in archive.list()
                    if item.is_file
                    and PurePosixPath(item.filename.replace("\\", "/")).suffix.lower()
                    in suffixes
                ]
                total_bytes = sum(int(item.uncompressed or 0) for item in selected)
                if total_bytes > max_total_bytes:
                    raise ValueError("FT压缩包解压后超过快速分析解压容量上限")
                destinations = [
                    _archive_destination(target, item.filename) for item in selected
                ]
                archive.extract(
                    path=target,
                    targets=[item.filename for item in selected],
                )
                extracted.extend(destinations)
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError("FT 7z压缩包损坏或无法读取") from exc
    else:
        raise ValueError("FT快速分析压缩包只支持ZIP或7z")
    if not extracted:
        expected = "、".join(sorted(suffixes))
        raise ValueError(f"FT压缩包内没有当前工具支持的源文件：{expected}")
    missing = [path for path in extracted if not path.is_file()]
    if missing:
        raise ValueError("FT压缩包未能完整解压")
    return tuple(sorted(extracted, key=lambda item: str(item).casefold()))


_FT_QUICK_PAT_SCRIPT = """
import os, sys
from pathlib import Path
package = os.environ['TMS_FT_PAT_PACKAGE']
source = Path(os.environ['TMS_FT_PAT_INPUT']).resolve()
output = Path(os.environ['TMS_FT_PAT_OUTPUT']).resolve()
adapter = os.environ['TMS_FT_PAT_ADAPTER']
sys.path.insert(0, package)
if adapter == 'JIEQUN_FT_QUICK_PAT_PYZ':
    from factories.jiequn.dc_auto import DC_FORMAT_UNIFIED, detect_dc_format
    from factories.jiequn.pat_cleaner import generate_raw_pat
    detection = detect_dc_format(source)
    if detection.format_name != DC_FORMAT_UNIFIED:
        raise SystemExit(
            f'Quick PAT only accepts Jiequn unified CSV, detected {detection.format_name}'
        )
    detected = {path.resolve() for path in detection.files}
    all_csv = {
        path.resolve() for path in source.rglob('*')
        if path.is_file() and path.suffix.lower() == '.csv'
    }
    if detected != all_csv:
        raise SystemExit(
            f'Jiequn unified CSV contract mismatch: detected={len(detected)}, csv={len(all_csv)}'
        )
elif adapter == 'RIYUEXIN_FT_QUICK_PAT_PYZ':
    from factories.riyuexin.pat_cleaner import generate_raw_pat
elif adapter == 'RIYUEGUANG_FT_QUICK_PAT_PYZ':
    from factories.tms_adapters.riyueguang_pat import generate_raw_pat
elif adapter == 'DIANJI_FT_QUICK_PAT_PYZ':
    from factories.dianji.pat_cleaner import generate_raw_pat
elif adapter == 'JIJIA_FT_QUICK_PAT_PYZ':
    from factories.jijia.pat_cleaner import generate_raw_pat
else:
    raise SystemExit(f'Unsupported FT Quick PAT adapter: {adapter}')
result = generate_raw_pat(source_dir=source, output_dir=output)
if not result:
    raise SystemExit('FT Quick PAT returned no result')
print(f'TMS_QUICK_PAT_RESULT={Path(result).resolve()}')
"""


_CP_QUICK_PAT_SCRIPT = """
import json, os, sys
sys.path.insert(0, os.environ['TMS_CP_PAT_PACKAGE'])
from cp_data_processor.analysis.quick_pat import generate_cleaned_csv_pat
result = generate_cleaned_csv_pat(
    cleaned_files=json.loads(os.environ['TMS_CP_PAT_CLEANED_JSON']),
    spec_files=json.loads(os.environ['TMS_CP_PAT_SPEC_JSON']),
    output_dir=os.environ['TMS_CP_PAT_OUTPUT'],
)
if not result:
    raise SystemExit('CP Quick PAT returned no result')
"""
