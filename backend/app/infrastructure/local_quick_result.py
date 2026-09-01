from __future__ import annotations

import hashlib
import json
import logging
import math
import os
import posixpath
import re
import shutil
import stat
import time
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO
from uuid import uuid4
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.errors import DomainError
from app.domain.cleaner_registry import CleanerRelease
from app.domain.quick_analysis import (
    LOCAL_QUICK_PAT_ADAPTER_CODE,
    LOCAL_QUICK_PAT_INPUT_CONTRACT,
    LOCAL_QUICK_PAT_OUTPUT_CONTRACT,
    LOCAL_QUICK_PAT_TOOL_CODE,
    LOCAL_RESULT_CONTRACT_VERSION,
    LocalQuickPatResultReceipt,
    QuickAnalysisArtifact,
)

_CHUNK_BYTES = 1024 * 1024
_DEFAULT_STAGING_TTL_SECONDS = 24 * 60 * 60
_MIN_STAGING_TTL_SECONDS = 60
_MAX_STAGING_TTL_SECONDS = 7 * 24 * 60 * 60
_XLSX_MAGIC = b"PK\x03\x04"
_SHA256_PATTERN = re.compile(r"^[0-9a-fA-F]{64}$")
_FORBIDDEN_XLSX_PARTS = (
    "xl/externalLinks/",
    "xl/embeddings/",
    "xl/oleObjects/",
)
_REQUIRED_PAT_XLSX_PARTS = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/app.xml",
        "docProps/core.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/styles.xml",
        "xl/theme/theme1.xml",
        "xl/workbook.xml",
        "xl/worksheets/sheet1.xml",
    }
)
_OPTIONAL_PAT_XLSX_PARTS = frozenset({"xl/sharedStrings.xml"})
_RELATIONSHIP_NAMESPACE = "http://schemas.openxmlformats.org/package/2006/relationships"
_OFFICE_RELATIONSHIP_PREFIX = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/"
)
_ROOT_RELATIONSHIPS = frozenset(
    {
        (
            f"{_OFFICE_RELATIONSHIP_PREFIX}officeDocument",
            "xl/workbook.xml",
        ),
        (
            f"{_RELATIONSHIP_NAMESPACE}/metadata/core-properties",
            "docProps/core.xml",
        ),
        (
            f"{_OFFICE_RELATIONSHIP_PREFIX}extended-properties",
            "docProps/app.xml",
        ),
    }
)
_WORKBOOK_RELATIONSHIPS = frozenset(
    {
        (f"{_OFFICE_RELATIONSHIP_PREFIX}worksheet", "xl/worksheets/sheet1.xml"),
        (f"{_OFFICE_RELATIONSHIP_PREFIX}styles", "xl/styles.xml"),
        (f"{_OFFICE_RELATIONSHIP_PREFIX}theme", "xl/theme/theme1.xml"),
    }
)
_PAT_HEADERS = (
    "统计量",
    "总计数",
    "均值",
    "标准差",
    "最小值",
    "下四分位数",
    "中位数",
    "上四分位数",
    "最大值",
    "Sigma",
    "LCL\n计算值",
    "UCL\n计算值",
    "LCL\n更新前",
    "UCL\n更新前",
    "LCL\n更新后",
    "UCL\n更新后",
    "是否\n更新",
)
logger = logging.getLogger(__name__)


@dataclass(frozen=True, slots=True)
class StagedLocalQuickResult:
    stage_dir: Path
    report_path: Path
    parameter_count: int


@dataclass(frozen=True, slots=True)
class CommittedLocalQuickResult:
    job_root: Path
    summary: dict[str, object]
    artifacts: tuple[QuickAnalysisArtifact, ...]


def validate_local_quick_pat_release(release: CleanerRelease) -> None:
    expected = {
        "test_stage": "FT",
        "factory_code": "JIEQUN",
        "format_code": LOCAL_QUICK_PAT_TOOL_CODE,
        "cleaner_code": LOCAL_QUICK_PAT_TOOL_CODE,
        "adapter_code": LOCAL_QUICK_PAT_ADAPTER_CODE,
        "input_contract_version": LOCAL_QUICK_PAT_INPUT_CONTRACT,
        "output_contract_version": LOCAL_QUICK_PAT_OUTPUT_CONTRACT,
    }
    mismatches = [
        name
        for name, value in expected.items()
        if str(getattr(release, name, "")).strip().upper() != value.upper()
    ]
    if mismatches:
        raise DomainError(
            "LOCAL_QUICK_CAPABILITY_UNAVAILABLE",
            "最新已发布 Cleaner 不是受批准的杰群 FT Quick PAT 版本",
            409,
            [{"mismatched_fields": mismatches}],
        )
    if not _SHA256_PATTERN.fullmatch(release.code_checksum or ""):
        raise DomainError(
            "LOCAL_QUICK_CAPABILITY_UNAVAILABLE",
            "Quick PAT Cleaner Release 缺少有效 SHA-256",
            409,
        )
    if release.max_output_bytes < 1 or release.timeout_seconds < 1:
        raise DomainError(
            "LOCAL_QUICK_CAPABILITY_UNAVAILABLE",
            "Quick PAT Cleaner Release 的执行限制无效",
            409,
        )


def local_quick_pat_capability(release: CleanerRelease) -> dict[str, object]:
    validate_local_quick_pat_release(release)
    return {
        "contract_version": LOCAL_RESULT_CONTRACT_VERSION,
        "tool_code": LOCAL_QUICK_PAT_TOOL_CODE,
        "analysis_type": "QUICK_PAT",
        "test_stage": "FT",
        "factory_code": "JIEQUN",
        "release": {
            "cleaner_release_id": release.cleaner_release_id,
            "format_profile_id": release.format_profile_id,
            "format_code": release.format_code,
            "profile_version": release.profile_version,
            "cleaner_code": release.cleaner_code,
            "cleaner_version": release.cleaner_version,
            "sha256": release.code_checksum.lower(),
            "entrypoint": release.entrypoint,
            "adapter_code": release.adapter_code,
            "input_contract_version": release.input_contract_version,
            "output_contract_version": release.output_contract_version,
            "timeout_seconds": release.timeout_seconds,
            "max_output_bytes": release.max_output_bytes,
        },
        "upload": {
            "multipart_receipt_field": "receipt_json",
            "multipart_result_field": "result_file",
            "accepted_extension": ".xlsx",
        },
    }


class LocalQuickResultStore:
    """Stage, validate, and atomically publish one Local Agent PAT result."""

    def __init__(
        self,
        work_root: str | Path,
        *,
        staging_ttl_seconds: int | None = None,
    ) -> None:
        self._work_root = Path(work_root).resolve()
        configured_ttl = (
            os.getenv(
                "TMS_LOCAL_RESULT_STAGING_TTL_SECONDS",
                str(_DEFAULT_STAGING_TTL_SECONDS),
            )
            if staging_ttl_seconds is None
            else str(staging_ttl_seconds)
        )
        try:
            self._staging_ttl_seconds = int(configured_ttl)
        except ValueError as exc:
            raise RuntimeError(
                "TMS_LOCAL_RESULT_STAGING_TTL_SECONDS must be an integer"
            ) from exc
        if not (
            _MIN_STAGING_TTL_SECONDS
            <= self._staging_ttl_seconds
            <= _MAX_STAGING_TTL_SECONDS
        ):
            raise RuntimeError(
                "TMS_LOCAL_RESULT_STAGING_TTL_SECONDS must be between "
                f"{_MIN_STAGING_TTL_SECONDS} and {_MAX_STAGING_TTL_SECONDS}"
            )

    @property
    def work_root(self) -> Path:
        return self._work_root

    def stage(
        self,
        stream: BinaryIO,
        *,
        upload_filename: str | None,
        receipt: LocalQuickPatResultReceipt,
        max_output_bytes: int,
    ) -> StagedLocalQuickResult:
        if receipt.result.size_bytes > max_output_bytes:
            raise DomainError(
                "LOCAL_RESULT_TOO_LARGE",
                "Local Agent PAT 结果超过 Cleaner Release 输出限制",
                413,
            )
        if upload_filename != receipt.result.filename:
            raise DomainError(
                "LOCAL_RESULT_FILENAME_MISMATCH",
                "上传文件名与 Local Agent 回执不一致",
                422,
            )
        self._work_root.mkdir(parents=True, exist_ok=True)
        staging_root = self._safe_staging_root()
        self.reap_stale_staging()
        stage_dir = staging_root / uuid4().hex
        stage_dir.mkdir()
        report = stage_dir / receipt.result.filename
        try:
            digest = hashlib.sha256()
            size_bytes = 0
            magic = bytearray()
            with report.open("xb") as target:
                while chunk := stream.read(_CHUNK_BYTES):
                    size_bytes += len(chunk)
                    if size_bytes > max_output_bytes:
                        raise DomainError(
                            "LOCAL_RESULT_TOO_LARGE",
                            "Local Agent PAT 结果超过 Cleaner Release 输出限制",
                            413,
                        )
                    if len(magic) < len(_XLSX_MAGIC):
                        magic.extend(chunk[: len(_XLSX_MAGIC) - len(magic)])
                    digest.update(chunk)
                    target.write(chunk)
            if size_bytes != receipt.result.size_bytes:
                raise DomainError(
                    "LOCAL_RESULT_SIZE_MISMATCH",
                    "上传 PAT 文件大小与 Local Agent 回执不一致",
                    422,
                )
            if bytes(magic) != _XLSX_MAGIC:
                raise DomainError(
                    "LOCAL_RESULT_XLSX_INVALID",
                    "上传结果不是有效 XLSX 文件",
                    422,
                )
            if digest.hexdigest() != receipt.result.sha256:
                raise DomainError(
                    "LOCAL_RESULT_SHA256_MISMATCH",
                    "上传 PAT 文件 SHA-256 与 Local Agent 回执不一致",
                    422,
                )
            self._preflight_xlsx(report, max_output_bytes)
            parameter_count = self._validate_pat_workbook(report, receipt)
            return StagedLocalQuickResult(stage_dir, report, parameter_count)
        except Exception:
            self.discard_staged(stage_dir)
            raise

    def commit(
        self,
        staged: StagedLocalQuickResult,
        *,
        job_id: int,
        receipt: LocalQuickPatResultReceipt,
        release: CleanerRelease,
    ) -> CommittedLocalQuickResult:
        if job_id < 1:
            raise ValueError("job_id must be positive")
        stage_dir = staged.stage_dir.resolve()
        if stage_dir.parent != self._safe_staging_root():
            raise RuntimeError("Local result staging directory escaped the work root")
        source_manifest = {
            "contract_version": LOCAL_RESULT_CONTRACT_VERSION,
            "source_label": receipt.source_label,
            **receipt.manifest.model_dump(mode="json"),
        }
        summary: dict[str, object] = {
            "contract_version": LOCAL_RESULT_CONTRACT_VERSION,
            "tool_code": receipt.tool_code,
            "analysis_type": receipt.analysis_type,
            "test_stage": receipt.test_stage,
            "factory_code": receipt.factory_code,
            "parameter_count": receipt.summary.parameter_count,
            "record_count": receipt.summary.record_count,
            "elapsed_seconds": receipt.summary.elapsed_seconds,
            "release": {
                "cleaner_release_id": release.cleaner_release_id,
                "cleaner_code": release.cleaner_code,
                "cleaner_version": release.cleaner_version,
                "sha256": release.code_checksum.lower(),
                "adapter_code": release.adapter_code,
                "input_contract_version": release.input_contract_version,
                "output_contract_version": release.output_contract_version,
            },
            "source_label": receipt.source_label,
            "manifest": receipt.manifest.model_dump(mode="json"),
            "summary": receipt.summary.model_dump(mode="json"),
            "result": receipt.result.model_dump(mode="json"),
        }
        (stage_dir / "source_manifest.json").write_text(
            json.dumps(source_manifest, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        (stage_dir / "pat_summary.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        job_root = (self._work_root / str(job_id)).resolve()
        if job_root.parent != self._work_root or job_root.name != str(job_id):
            raise RuntimeError("Local result job directory escaped the work root")
        if job_root.exists():
            raise DomainError(
                "LOCAL_RESULT_TARGET_EXISTS",
                "Local Agent 结果任务目录已存在",
                409,
            )
        try:
            os.replace(stage_dir, job_root)
            artifacts = tuple(
                self._artifact(role, job_root / filename)
                for role, filename in (
                    ("pat_report", receipt.result.filename),
                    ("pat_summary", "pat_summary.json"),
                    ("source_manifest", "source_manifest.json"),
                )
            )
            return CommittedLocalQuickResult(job_root, summary, artifacts)
        except Exception:
            if job_root.exists():
                self.discard_committed(job_root)
            raise

    def discard_staged(self, stage_dir: str | Path) -> None:
        unresolved = Path(stage_dir)
        if _is_link_or_reparse_point(unresolved):
            raise RuntimeError("refusing to remove a linked Local result staging path")
        candidate = unresolved.resolve()
        staging_root = self._safe_staging_root()
        if candidate.parent != staging_root or not re.fullmatch(
            r"[0-9a-f]{32}", candidate.name
        ):
            raise RuntimeError("refusing to remove an unsafe Local result staging path")
        if candidate.exists():
            self._remove_staging_tree(candidate)

    def reap_stale_staging(
        self,
        *,
        now_epoch_seconds: float | None = None,
    ) -> tuple[str, ...]:
        """Remove only stale, link-free UUID staging directories."""
        staging_root = self._safe_staging_root()
        cutoff = (
            time.time() if now_epoch_seconds is None else now_epoch_seconds
        ) - self._staging_ttl_seconds
        removed: list[str] = []
        with os.scandir(staging_root) as entries:
            candidates = sorted(entries, key=lambda entry: entry.name)
        for entry in candidates:
            if not re.fullmatch(r"[0-9a-f]{32}", entry.name):
                continue
            candidate = staging_root / entry.name
            try:
                if (
                    entry.is_symlink()
                    or _is_link_or_reparse_point(candidate)
                    or not entry.is_dir(follow_symlinks=False)
                    or entry.stat(follow_symlinks=False).st_mtime > cutoff
                ):
                    continue
                if _tree_contains_link_or_reparse_point(candidate):
                    logger.warning(
                        "refusing to reap linked Local result staging tree %s",
                        candidate,
                    )
                    continue
                self._remove_staging_tree(candidate)
                removed.append(entry.name)
            except FileNotFoundError:
                continue
        return tuple(removed)

    def discard_committed(self, job_root: str | Path) -> None:
        candidate = Path(job_root).resolve()
        if (
            candidate.parent != self._work_root
            or not candidate.name.isdigit()
            or int(candidate.name) < 1
        ):
            raise RuntimeError("refusing to remove an unsafe Local result job path")
        if candidate.exists():
            shutil.rmtree(candidate)

    def _safe_staging_root(self) -> Path:
        staging = self._work_root / ".staging"
        if _is_link_or_reparse_point(staging):
            raise RuntimeError("Local result staging root must not be a link")
        staging.mkdir(parents=True, exist_ok=True)
        if _is_link_or_reparse_point(staging):
            raise RuntimeError("Local result staging root must not be a link")
        resolved = staging.resolve()
        if resolved.parent != self._work_root or resolved.name != ".staging":
            raise RuntimeError("Local result staging root escaped the work root")
        return resolved

    def _remove_staging_tree(self, candidate: Path) -> None:
        if (
            candidate.parent != self._safe_staging_root()
            or not re.fullmatch(r"[0-9a-f]{32}", candidate.name)
            or _is_link_or_reparse_point(candidate)
            or _tree_contains_link_or_reparse_point(candidate)
        ):
            raise RuntimeError("refusing to remove an unsafe Local result staging tree")
        shutil.rmtree(candidate)

    @staticmethod
    def _preflight_xlsx(path: Path, max_output_bytes: int) -> None:
        expanded_limit = min(max_output_bytes * 20, 256 * 1024**2)
        try:
            with ZipFile(path) as archive:
                entries = archive.infolist()
                names = {entry.filename for entry in entries}
                if (
                    not entries
                    or len(entries) > 1000
                    or len(names) != len(entries)
                    or any(entry.flag_bits & 0x1 for entry in entries)
                    or sum(entry.file_size for entry in entries) > expanded_limit
                    or "xl/workbook.xml" not in names
                ):
                    raise DomainError(
                        "LOCAL_RESULT_XLSX_INVALID",
                        "上传 XLSX 的容器结构不符合 PAT 结果合同",
                        422,
                    )
                allowed_parts = _REQUIRED_PAT_XLSX_PARTS | _OPTIONAL_PAT_XLSX_PARTS
                if (
                    not _REQUIRED_PAT_XLSX_PARTS.issubset(names)
                    or not names.issubset(allowed_parts)
                ):
                    raise DomainError(
                        "LOCAL_RESULT_XLSX_PARTS_FORBIDDEN",
                        "PAT 结果包含输出合同未批准的 XLSX 部件",
                        422,
                    )
                _validate_xlsx_relationships(archive, names)
                if (
                    "xl/connections.xml" in names
                    or "xl/vbaProject.bin" in names
                    or any(
                        name.startswith(prefix)
                        for name in names
                        for prefix in _FORBIDDEN_XLSX_PARTS
                    )
                    or _xlsx_xml_has_active_content(archive, names)
                ):
                    raise DomainError(
                        "LOCAL_RESULT_ACTIVE_CONTENT_FORBIDDEN",
                        "PAT 结果不得包含公式、外部连接或嵌入对象",
                        422,
                    )
        except BadZipFile as exc:
            raise DomainError(
                "LOCAL_RESULT_XLSX_INVALID", "上传结果不是有效 XLSX 文件", 422
            ) from exc

    @staticmethod
    def _validate_pat_workbook(
        path: Path, receipt: LocalQuickPatResultReceipt
    ) -> int:
        try:
            workbook = load_workbook(path, read_only=True, data_only=False)
        except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as exc:
            raise DomainError(
                "LOCAL_RESULT_XLSX_INVALID", "上传结果不是可读的 PAT XLSX", 422
            ) from exc
        rows = None
        validation_error: tuple[str, str, int, list[dict]] | None = None
        parameter_count = 0
        try:
            if workbook.sheetnames != ["PAT"]:
                raise DomainError(
                    "LOCAL_RESULT_PAT_SHEET_INVALID",
                    "PAT 结果必须且只能包含名为 PAT 的工作表",
                    422,
                )
            if getattr(workbook, "_external_links", ()) or len(workbook.defined_names):
                raise DomainError(
                    "LOCAL_RESULT_ACTIVE_CONTENT_FORBIDDEN",
                    "PAT 结果不得包含外部链接或定义名称",
                    422,
                )
            sheet = workbook["PAT"]
            if sheet.max_column != len(_PAT_HEADERS):
                raise DomainError(
                    "LOCAL_RESULT_PAT_HEADER_INVALID",
                    "PAT 工作表列数与 FT_PAT_RESULT_V1 不一致",
                    422,
                )
            rows = sheet.iter_rows()
            first = _safe_pat_row(next(rows, ()))
            second = _safe_pat_row(next(rows, ()))
            expected_second = ("变量", *_PAT_HEADERS[1:])
            if first != _PAT_HEADERS or second != expected_second:
                raise DomainError(
                    "LOCAL_RESULT_PAT_HEADER_INVALID",
                    "PAT 工作表表头与 FT_PAT_RESULT_V1 不一致",
                    422,
                )
            parameters: set[str] = set()
            for cells in rows:
                normalized = _safe_pat_row(cells)
                if not any(value is not None and str(value).strip() for value in normalized):
                    continue
                parameter = str(normalized[0] or "").strip()
                count = normalized[1] if len(normalized) > 1 else None
                numeric_statistics = normalized[1:16]
                if (
                    not parameter
                    or len(parameter) > 200
                    or parameter in parameters
                    or isinstance(count, bool)
                    or not isinstance(count, (int, float))
                    or not math.isfinite(float(count))
                    or float(count) < 1
                    or not float(count).is_integer()
                    or int(count) > receipt.summary.record_count
                    or any(
                        value is not None
                        and (
                            isinstance(value, bool)
                            or not isinstance(value, (int, float))
                            or not math.isfinite(float(value))
                        )
                        for value in numeric_statistics
                    )
                ):
                    raise DomainError(
                        "LOCAL_RESULT_PAT_ROWS_INVALID",
                        "PAT 工作表参数行与回执摘要不一致",
                        422,
                    )
                parameters.add(parameter)
            parameter_count = len(parameters)
            if parameter_count != receipt.summary.parameter_count:
                raise DomainError(
                    "LOCAL_RESULT_PARAMETER_COUNT_MISMATCH",
                    "PAT 工作表参数数与 Local Agent 回执不一致",
                    422,
                )
        except DomainError as exc:
            validation_error = (
                exc.code,
                exc.message,
                exc.status_code,
                exc.details,
            )
        finally:
            if rows is not None:
                rows.close()
            workbook.close()
        if validation_error is not None:
            raise DomainError(*validation_error)
        return parameter_count

    @staticmethod
    def _artifact(role: str, path: Path) -> QuickAnalysisArtifact:
        resolved = path.resolve()
        return QuickAnalysisArtifact(
            role,
            str(resolved),
            resolved.stat().st_size,
            _file_sha256(resolved),
        )


def _safe_pat_row(cells) -> tuple[object, ...]:
    values: list[object] = []
    for cell in tuple(cells)[: len(_PAT_HEADERS)]:
        if getattr(cell, "data_type", None) == "f" or getattr(cell, "hyperlink", None):
            raise DomainError(
                "LOCAL_RESULT_ACTIVE_CONTENT_FORBIDDEN",
                "PAT 结果不得包含公式或超链接",
                422,
            )
        values.append(getattr(cell, "value", None))
    return tuple(values)


def _xlsx_xml_has_active_content(archive: ZipFile, names: set[str]) -> bool:
    worksheet_parts = (
        name
        for name in names
        if name.startswith("xl/worksheets/") and name.endswith(".xml")
    )
    active_pattern = re.compile(br"<(?:[A-Za-z0-9_]+:)?(?:f|hyperlink)(?:\s|>)")
    if any(active_pattern.search(archive.read(name)) for name in worksheet_parts):
        return True
    workbook_xml = archive.read("xl/workbook.xml")
    return bool(
        re.search(
            br"<(?:[A-Za-z0-9_]+:)?definedName(?:\s|>)",
            workbook_xml,
        )
    )


def _validate_xlsx_relationships(archive: ZipFile, names: set[str]) -> None:
    root_relationships = _read_relationships(
        archive,
        "_rels/.rels",
        base_directory="",
    )
    if root_relationships != _ROOT_RELATIONSHIPS:
        raise DomainError(
            "LOCAL_RESULT_XLSX_RELATIONSHIPS_FORBIDDEN",
            "PAT 结果包含输出合同未批准的根关系",
            422,
        )

    expected_workbook_relationships = set(_WORKBOOK_RELATIONSHIPS)
    if "xl/sharedStrings.xml" in names:
        expected_workbook_relationships.add(
            (
                f"{_OFFICE_RELATIONSHIP_PREFIX}sharedStrings",
                "xl/sharedStrings.xml",
            )
        )
    workbook_relationships = _read_relationships(
        archive,
        "xl/_rels/workbook.xml.rels",
        base_directory="xl",
    )
    if workbook_relationships != frozenset(expected_workbook_relationships):
        raise DomainError(
            "LOCAL_RESULT_XLSX_RELATIONSHIPS_FORBIDDEN",
            "PAT 结果包含输出合同未批准的工作簿关系",
            422,
        )


def _read_relationships(
    archive: ZipFile,
    part_name: str,
    *,
    base_directory: str,
) -> frozenset[tuple[str, str]]:
    try:
        root = ET.fromstring(archive.read(part_name))
    except (ET.ParseError, KeyError) as exc:
        raise DomainError(
            "LOCAL_RESULT_XLSX_INVALID",
            "PAT 结果的 XLSX 关系结构无效",
            422,
        ) from exc
    relationship_tag = f"{{{_RELATIONSHIP_NAMESPACE}}}Relationship"
    relationships: set[tuple[str, str]] = set()
    identifiers: set[str] = set()
    if root.tag != f"{{{_RELATIONSHIP_NAMESPACE}}}Relationships":
        raise DomainError(
            "LOCAL_RESULT_XLSX_INVALID",
            "PAT 结果的 XLSX 关系结构无效",
            422,
        )
    for child in root:
        identifier = child.attrib.get("Id", "")
        relationship_type = child.attrib.get("Type", "")
        target = child.attrib.get("Target", "")
        if (
            child.tag != relationship_tag
            or not identifier
            or identifier in identifiers
            or child.attrib.get("TargetMode") is not None
            or not relationship_type
            or not target
            or "\\" in target
        ):
            raise DomainError(
                "LOCAL_RESULT_XLSX_RELATIONSHIPS_FORBIDDEN",
                "PAT 结果包含输出合同未批准的 XLSX 关系",
                422,
            )
        identifiers.add(identifier)
        normalized_target = posixpath.normpath(
            target.lstrip("/")
            if target.startswith("/")
            else posixpath.join(base_directory, target)
        )
        if normalized_target == ".." or normalized_target.startswith("../"):
            raise DomainError(
                "LOCAL_RESULT_XLSX_RELATIONSHIPS_FORBIDDEN",
                "PAT 结果包含越界的 XLSX 关系",
                422,
            )
        relationships.add((relationship_type, normalized_target))
    if len(relationships) != len(identifiers):
        raise DomainError(
            "LOCAL_RESULT_XLSX_RELATIONSHIPS_FORBIDDEN",
            "PAT 结果包含重复的 XLSX 关系",
            422,
        )
    return frozenset(relationships)


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(_CHUNK_BYTES):
            digest.update(chunk)
    return digest.hexdigest()


def _is_link_or_reparse_point(path: Path) -> bool:
    try:
        if path.is_symlink():
            return True
        is_junction = getattr(path, "is_junction", None)
        if is_junction is not None and is_junction():
            return True
        attributes = getattr(path.lstat(), "st_file_attributes", 0)
    except FileNotFoundError:
        return False
    reparse_flag = getattr(stat, "FILE_ATTRIBUTE_REPARSE_POINT", 0x400)
    return bool(attributes & reparse_flag)


def _tree_contains_link_or_reparse_point(root: Path) -> bool:
    pending = [root]
    while pending:
        current = pending.pop()
        try:
            with os.scandir(current) as entries:
                children = list(entries)
        except FileNotFoundError:
            return True
        for entry in children:
            candidate = Path(entry.path)
            try:
                if entry.is_symlink() or _is_link_or_reparse_point(candidate):
                    return True
                if entry.is_dir(follow_symlinks=False):
                    pending.append(candidate)
            except FileNotFoundError:
                return True
    return False
