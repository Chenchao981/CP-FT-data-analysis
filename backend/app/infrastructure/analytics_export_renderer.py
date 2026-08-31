from __future__ import annotations

import csv
import html
import io
import json
import logging
import os
from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:  # pragma: no cover - exercised through dependency sentinel tests
    Image = ImageDraw = ImageFont = None  # type: ignore[assignment]

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas as pdf_canvas
except ImportError:  # pragma: no cover - exercised through dependency sentinel tests
    A4 = landscape = pdfmetrics = TTFont = pdf_canvas = None  # type: ignore[assignment]

from app.core.errors import DomainError
from app.domain.analytics_export_worker import (
    AnalyticsExportContentSource,
    AnalyticsExportWorkItem,
    ExportCell,
    RenderedAnalyticsExport,
)
from app.domain.saved_analyses import canonical_json
from app.infrastructure.analytics_export_files import AnalyticsExportPathPolicy

_LOGGER = logging.getLogger(__name__)

_FORMAT_CONTRACT = {
    "CSV": ("csv", "text/csv; charset=utf-8"),
    "XLSX": (
        "xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ),
    "BIN_TXT": ("txt", "text/plain"),
    "HTML": ("html", "text/html; charset=utf-8"),
    "PDF": ("pdf", "application/pdf"),
    "PNG": ("png", "image/png"),
}
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")
_XLSX_CONTEXT_CHUNK_SIZE = 30_000
_REPORT_PREVIEW_ROWS = 50
_PDF_FONT_NAME = "TMSAnalyticsUnicode"


def _safe_cell(value: ExportCell) -> ExportCell:
    if isinstance(value, str):
        cleaned = value.replace("\x00", "")
        if cleaned.startswith(_FORMULA_PREFIXES):
            return "'" + cleaned
        return cleaned
    return value


def _text_cell(value: ExportCell) -> str:
    safe = _safe_cell(value)
    if safe is None:
        return ""
    if isinstance(safe, bool):
        return "TRUE" if safe else "FALSE"
    return str(safe)


def _font_candidates() -> tuple[Path, ...]:
    windows = Path(os.environ.get("WINDIR", r"C:\Windows")) / "Fonts"
    return (
        windows / "msyh.ttc",
        windows / "simhei.ttf",
        windows / "simsun.ttc",
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
        Path("/System/Library/Fonts/PingFang.ttc"),
    )


def _ascii_fallback(value: str, *, unicode_font: bool) -> str:
    if unicode_font:
        return value
    return value.encode("ascii", "backslashreplace").decode("ascii")


def _clip_characters(value: str, limit: int) -> str:
    if len(value) <= limit:
        return value
    return value[: max(1, limit - 14)] + "...[truncated]"


class AnalyticsExportRenderer:
    def __init__(
        self,
        path_policy: AnalyticsExportPathPolicy,
        content_source: AnalyticsExportContentSource,
    ) -> None:
        self._path_policy = path_policy
        self._content_source = content_source

    def render(self, work_item: AnalyticsExportWorkItem) -> RenderedAnalyticsExport:
        format_name = work_item.export_format.value
        contract = _FORMAT_CONTRACT.get(format_name)
        if contract is None:
            raise DomainError(
                "ANALYTICS_EXPORT_FORMAT_UNSUPPORTED",
                "the requested analytics export format is unsupported",
                409,
            )
        self._require_format_dependency(format_name)
        extension, mime_type = contract
        file_name = (
            f"analytics-export-{work_item.export_job_id}-"
            f"attempt-{work_item.attempt_count}.{extension}"
        )
        self._path_policy.prepare_job_root(work_item.export_job_id)
        try:
            table = self._content_source.table(work_item)
            if format_name == "CSV":
                row_count = self._write_delimited(
                    work_item,
                    file_name,
                    table.columns,
                    table.rows,
                    delimiter=",",
                    bom=True,
                )
            elif format_name == "BIN_TXT":
                row_count = self._write_delimited(
                    work_item,
                    file_name,
                    table.columns,
                    table.rows,
                    delimiter="\t",
                    bom=False,
                )
            elif format_name == "XLSX":
                row_count = self._write_xlsx(
                    work_item, file_name, table.columns, table.rows
                )
            elif format_name == "HTML":
                row_count = self._write_html(
                    work_item, file_name, table.columns, table.rows
                )
            elif format_name == "PDF":
                row_count = self._write_pdf(
                    work_item, file_name, table.columns, table.rows
                )
            else:
                row_count = self._write_png(
                    work_item, file_name, table.columns, table.rows
                )
        except Exception:
            self._path_policy.remove_empty_job_root(work_item.export_job_id)
            raise
        identity = self._path_policy.identify(
            work_item.export_job_id,
            self._path_policy.artifact_path(work_item.export_job_id, file_name),
        )
        return RenderedAnalyticsExport(
            path=identity.path,
            file_name=identity.file_name,
            mime_type=mime_type,
            file_size=identity.file_size,
            sha256=identity.sha256,
            exported_row_count=row_count,
        )

    def discard(
        self,
        work_item: AnalyticsExportWorkItem,
        artifact: RenderedAnalyticsExport,
    ) -> None:
        """Discard only this fenced attempt's unregistered output."""

        self._path_policy.remove_artifact(work_item.export_job_id, artifact.path)

    @staticmethod
    def _require_format_dependency(format_name: str) -> None:
        available = {
            "PDF": pdf_canvas is not None and pdfmetrics is not None,
            "PNG": Image is not None
            and ImageDraw is not None
            and ImageFont is not None,
        }
        if format_name in available and not available[format_name]:
            raise DomainError(
                "ANALYTICS_EXPORT_DEPENDENCY_UNAVAILABLE",
                f"{format_name} rendering dependency is unavailable",
                409,
            )

    def _write_delimited(
        self,
        work_item: AnalyticsExportWorkItem,
        file_name: str,
        columns: tuple[str, ...],
        rows: Iterable[tuple[ExportCell, ...]],
        *,
        delimiter: str,
        bom: bool,
    ) -> int:
        row_count = 0
        with self._path_policy.atomic_binary_writer(
            work_item.export_job_id, file_name
        ) as (binary, _temporary):
            if bom:
                binary.write(b"\xef\xbb\xbf")
            text_stream = io.TextIOWrapper(
                binary, encoding="utf-8", newline="", write_through=True
            )
            try:
                writer = csv.writer(
                    text_stream,
                    delimiter=delimiter,
                    lineterminator="\r\n",
                    quoting=csv.QUOTE_MINIMAL,
                )
                writer.writerow(_safe_cell(column) for column in columns)
                for row in rows:
                    self._require_width(columns, row)
                    writer.writerow(_safe_cell(value) for value in row)
                    row_count += 1
                text_stream.flush()
            finally:
                text_stream.detach()
        return row_count

    @staticmethod
    def _context_rows(
        work_item: AnalyticsExportWorkItem,
    ) -> tuple[tuple[str, str], ...]:
        return (
            ("contract", "ANALYTICS_EXPORT_WORKER_V1"),
            ("export_job_id", str(work_item.export_job_id)),
            ("template", f"{work_item.template_code}:{work_item.template_version}"),
            ("scope", work_item.export_scope.value),
            ("format", work_item.export_format.value),
            ("test_stage", work_item.test_stage),
            ("filter_hash", work_item.filter_hash),
            ("context_hash", work_item.context_hash),
            ("presentation_hash", work_item.presentation_hash),
            (
                "datasets",
                json.dumps(
                    [
                        item.model_dump(mode="json")
                        for item in work_item.context.datasets
                    ],
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
            ),
            (
                "filters",
                work_item.context.filters.model_dump_json(exclude_none=True),
            ),
            (
                "parameters",
                json.dumps(
                    work_item.context.parameters,
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
            ),
            (
                "rule_context",
                work_item.rule_context.model_dump_json(exclude_none=True),
            ),
            ("chart_config", canonical_json(work_item.chart_config)),
            ("display_config", canonical_json(work_item.display_config)),
        )

    def _write_xlsx(
        self,
        work_item: AnalyticsExportWorkItem,
        file_name: str,
        columns: tuple[str, ...],
        rows: Iterable[tuple[ExportCell, ...]],
    ) -> int:
        if len(columns) > 16_384:
            raise DomainError(
                "ANALYTICS_EXPORT_XLSX_COLUMN_LIMIT",
                "the export exceeds the XLSX column limit",
                409,
            )
        workbook = Workbook(write_only=True)
        context_sheet = workbook.create_sheet("Context")
        context_sheet.append(("Field", "Value"))
        for key, value in self._context_rows(work_item):
            chunks = tuple(
                value[offset : offset + _XLSX_CONTEXT_CHUNK_SIZE]
                for offset in range(0, len(value), _XLSX_CONTEXT_CHUNK_SIZE)
            ) or ("",)
            for index, chunk in enumerate(chunks, start=1):
                label = key if len(chunks) == 1 else f"{key}[{index}/{len(chunks)}]"
                context_sheet.append((label, _safe_cell(chunk)))

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        def new_data_sheet(index: int):
            name = "Data" if index == 1 else f"Data_{index}"
            sheet = workbook.create_sheet(name)
            header = []
            for column in columns:
                cell = WriteOnlyCell(sheet, value=_safe_cell(column))
                cell.fill = header_fill
                cell.font = header_font
                header.append(cell)
            sheet.append(header)
            sheet.freeze_panes = "A2"
            return sheet

        sheet_index = 1
        row_in_sheet = 1
        sheet = new_data_sheet(sheet_index)
        row_count = 0
        for row in rows:
            self._require_width(columns, row)
            if row_in_sheet >= 1_048_576:
                sheet_index += 1
                sheet = new_data_sheet(sheet_index)
                row_in_sheet = 1
            sheet.append(tuple(_safe_cell(value) for value in row))
            row_count += 1
            row_in_sheet += 1

        with self._path_policy.atomic_binary_writer(
            work_item.export_job_id, file_name
        ) as (binary, _temporary):
            workbook.save(binary)
        workbook.close()
        return row_count

    def _write_html(
        self,
        work_item: AnalyticsExportWorkItem,
        file_name: str,
        columns: tuple[str, ...],
        rows: Iterable[tuple[ExportCell, ...]],
    ) -> int:
        row_count = 0
        with self._path_policy.atomic_binary_writer(
            work_item.export_job_id, file_name
        ) as (binary, _temporary):
            text_stream = io.TextIOWrapper(
                binary, encoding="utf-8", newline="", write_through=True
            )
            try:
                text_stream.write(
                    '<!doctype html><html lang="zh-CN"><head><meta charset="utf-8">'
                    '<meta name="viewport" content="width=device-width,initial-scale=1">'
                    "<title>TMS Analytics Export</title><style>"
                    "body{font-family:Segoe UI,Arial,sans-serif;margin:24px;color:#1f2937}"
                    "h1{font-size:22px}table{border-collapse:collapse;width:100%;font-size:12px}"
                    "th,td{border:1px solid #d1d5db;padding:6px;text-align:left}"
                    "th{background:#1f4e78;color:white;position:sticky;top:0}"
                    ".context{margin-bottom:20px}.context th{position:static;width:180px}"
                    "</style></head><body><h1>TMS Analytics Export</h1>"
                    '<table class="context"><tbody>'
                )
                for key, value in self._context_rows(work_item):
                    text_stream.write(
                        f"<tr><th>{html.escape(key)}</th><td>{html.escape(value)}</td></tr>"
                    )
                text_stream.write("</tbody></table><table><thead><tr>")
                for column in columns:
                    text_stream.write(f"<th>{html.escape(column)}</th>")
                text_stream.write("</tr></thead><tbody>")
                for row in rows:
                    self._require_width(columns, row)
                    text_stream.write("<tr>")
                    for value in row:
                        text_stream.write(f"<td>{html.escape(_text_cell(value))}</td>")
                    text_stream.write("</tr>")
                    row_count += 1
                text_stream.write("</tbody></table></body></html>")
                text_stream.flush()
            finally:
                text_stream.detach()
        return row_count

    @classmethod
    def _collect_report_preview(
        cls,
        columns: tuple[str, ...],
        rows: Iterable[tuple[ExportCell, ...]],
    ) -> tuple[int, tuple[tuple[str, ...], ...]]:
        if not columns:
            raise DomainError(
                "ANALYTICS_EXPORT_CONTENT_WIDTH_INVALID",
                "the server report renderer requires at least one column",
                409,
            )
        row_count = 0
        preview: list[tuple[str, ...]] = []
        for row in rows:
            cls._require_width(columns, row)
            if len(preview) < _REPORT_PREVIEW_ROWS:
                preview.append(tuple(_text_cell(value) for value in row))
            row_count += 1
        return row_count, tuple(preview)

    @staticmethod
    def _pdf_font() -> tuple[str, bool]:
        assert pdfmetrics is not None and TTFont is not None
        for candidate in _font_candidates():
            if not candidate.is_file():
                continue
            try:
                if _PDF_FONT_NAME not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(_PDF_FONT_NAME, str(candidate)))
                return _PDF_FONT_NAME, True
            except Exception as exc:  # noqa: BLE001 - font parsers vary by platform
                _LOGGER.debug(
                    "analytics PDF font candidate rejected exception_type=%s",
                    type(exc).__name__,
                )
                continue
        return "Helvetica", False

    @staticmethod
    def _pdf_fit(value: str, *, font_name: str, font_size: float, width: float) -> str:
        assert pdfmetrics is not None
        if pdfmetrics.stringWidth(value, font_name, font_size) <= width:
            return value
        suffix = "..."
        low = 0
        high = len(value)
        while low < high:
            middle = (low + high + 1) // 2
            candidate = value[:middle] + suffix
            if pdfmetrics.stringWidth(candidate, font_name, font_size) <= width:
                low = middle
            else:
                high = middle - 1
        return value[:low] + suffix

    def _write_pdf(
        self,
        work_item: AnalyticsExportWorkItem,
        file_name: str,
        columns: tuple[str, ...],
        rows: Iterable[tuple[ExportCell, ...]],
    ) -> int:
        assert pdf_canvas is not None and landscape is not None and A4 is not None
        row_count, preview = self._collect_report_preview(columns, rows)
        font_name, unicode_font = self._pdf_font()
        page_size = landscape(A4)
        page_width, page_height = page_size
        margin = 28.0
        context_font_size = 8.5
        table_font_size = 7.5
        row_height = 17.0

        with self._path_policy.atomic_binary_writer(
            work_item.export_job_id, file_name
        ) as (binary, _temporary):
            document = pdf_canvas.Canvas(binary, pagesize=page_size, pageCompression=1)
            document.setTitle("TMS Analytics Export")
            document.setAuthor("TMS Analytics Export Worker")
            document.setSubject(
                f"{work_item.template_code}@{work_item.template_version}"
            )
            document.setFont(font_name, 16)
            document.drawString(margin, page_height - margin, "TMS Analytics Export")
            y = page_height - margin - 24
            document.setFont(font_name, context_font_size)
            for key, raw_value in self._context_rows(work_item):
                value = _clip_characters(raw_value, 320)
                display = _ascii_fallback(f"{key}: {value}", unicode_font=unicode_font)
                display = self._pdf_fit(
                    display,
                    font_name=font_name,
                    font_size=context_font_size,
                    width=page_width - 2 * margin,
                )
                document.drawString(margin, y, display)
                y -= 12
            font_mode = (
                "Unicode system font"
                if unicode_font
                else "ASCII backslash-escape fallback"
            )
            truncated = row_count > len(preview)
            report_note = (
                f"source_rows={row_count}; shown_rows={len(preview)}; "
                f"preview_truncated={'YES' if truncated else 'NO'}; font={font_mode}"
            )
            document.drawString(margin, y, report_note)
            y -= 20

            column_width = (page_width - 2 * margin) / len(columns)

            def draw_header(top: float) -> float:
                document.setFillColorRGB(0.12, 0.31, 0.47)
                document.rect(
                    margin,
                    top - row_height,
                    page_width - 2 * margin,
                    row_height,
                    stroke=0,
                    fill=1,
                )
                document.setFillColorRGB(1, 1, 1)
                document.setFont(font_name, table_font_size)
                for index, column in enumerate(columns):
                    value = _ascii_fallback(str(column), unicode_font=unicode_font)
                    value = self._pdf_fit(
                        value,
                        font_name=font_name,
                        font_size=table_font_size,
                        width=column_width - 6,
                    )
                    document.drawString(
                        margin + index * column_width + 3,
                        top - row_height + 5,
                        value,
                    )
                document.setFillColorRGB(0.12, 0.15, 0.18)
                return top - row_height

            y = draw_header(y)
            for row in preview:
                if y - row_height < margin:
                    document.showPage()
                    document.setFont(font_name, 11)
                    document.drawString(
                        margin,
                        page_height - margin,
                        "TMS Analytics Export - table preview (continued)",
                    )
                    y = draw_header(page_height - margin - 20)
                document.setStrokeColorRGB(0.82, 0.84, 0.87)
                document.setFont(font_name, table_font_size)
                for index, raw_value in enumerate(row):
                    left = margin + index * column_width
                    document.rect(
                        left,
                        y - row_height,
                        column_width,
                        row_height,
                        stroke=1,
                        fill=0,
                    )
                    value = _ascii_fallback(raw_value, unicode_font=unicode_font)
                    value = self._pdf_fit(
                        value,
                        font_name=font_name,
                        font_size=table_font_size,
                        width=column_width - 6,
                    )
                    document.drawString(left + 3, y - row_height + 5, value)
                y -= row_height
            document.save()
        return row_count

    @staticmethod
    def _pillow_font(size: int):
        assert ImageFont is not None
        for candidate in _font_candidates():
            if not candidate.is_file():
                continue
            try:
                return ImageFont.truetype(str(candidate), size=size), True
            except (OSError, ValueError):
                continue
        return ImageFont.load_default(), False

    @staticmethod
    def _pillow_fit(draw, value: str, *, font, width: int) -> str:
        if draw.textbbox((0, 0), value, font=font)[2] <= width:
            return value
        suffix = "..."
        low = 0
        high = len(value)
        while low < high:
            middle = (low + high + 1) // 2
            candidate = value[:middle] + suffix
            if draw.textbbox((0, 0), candidate, font=font)[2] <= width:
                low = middle
            else:
                high = middle - 1
        return value[:low] + suffix

    def _write_png(
        self,
        work_item: AnalyticsExportWorkItem,
        file_name: str,
        columns: tuple[str, ...],
        rows: Iterable[tuple[ExportCell, ...]],
    ) -> int:
        assert Image is not None and ImageDraw is not None
        row_count, preview = self._collect_report_preview(columns, rows)
        title_font, title_unicode = self._pillow_font(28)
        body_font, body_unicode = self._pillow_font(16)
        unicode_font = title_unicode and body_unicode
        width = 2_000
        margin = 32
        context_row_height = 24
        table_row_height = 34
        context_rows = self._context_rows(work_item)
        height = (
            margin
            + 44
            + len(context_rows) * context_row_height
            + 42
            + (len(preview) + 1) * table_row_height
            + margin
        )
        image = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(image)
        draw.text(
            (margin, margin), "TMS Analytics Export", fill="#111827", font=title_font
        )
        y = margin + 48
        for key, raw_value in context_rows:
            value = _ascii_fallback(
                f"{key}: {_clip_characters(raw_value, 320)}",
                unicode_font=unicode_font,
            )
            value = self._pillow_fit(
                draw, value, font=body_font, width=width - 2 * margin
            )
            draw.text((margin, y), value, fill="#374151", font=body_font)
            y += context_row_height
        font_mode = (
            "Unicode system font" if unicode_font else "ASCII backslash-escape fallback"
        )
        note = (
            f"source_rows={row_count}; shown_rows={len(preview)}; "
            f"preview_truncated={'YES' if row_count > len(preview) else 'NO'}; "
            f"font={font_mode}"
        )
        draw.text((margin, y + 4), note, fill="#111827", font=body_font)
        y += 42

        column_width = (width - 2 * margin) / len(columns)
        draw.rectangle(
            (margin, y, width - margin, y + table_row_height),
            fill="#1f4e78",
        )
        for index, column in enumerate(columns):
            left = int(margin + index * column_width)
            right = int(margin + (index + 1) * column_width)
            value = _ascii_fallback(str(column), unicode_font=unicode_font)
            value = self._pillow_fit(
                draw, value, font=body_font, width=max(1, right - left - 8)
            )
            draw.text((left + 4, y + 7), value, fill="white", font=body_font)
        y += table_row_height
        for row in preview:
            for index, raw_value in enumerate(row):
                left = int(margin + index * column_width)
                right = int(margin + (index + 1) * column_width)
                draw.rectangle(
                    (left, y, right, y + table_row_height),
                    outline="#d1d5db",
                    width=1,
                )
                value = _ascii_fallback(raw_value, unicode_font=unicode_font)
                value = self._pillow_fit(
                    draw, value, font=body_font, width=max(1, right - left - 8)
                )
                draw.text((left + 4, y + 7), value, fill="#1f2937", font=body_font)
            y += table_row_height

        try:
            with self._path_policy.atomic_binary_writer(
                work_item.export_job_id, file_name
            ) as (binary, _temporary):
                image.save(binary, format="PNG", optimize=True)
        finally:
            image.close()
        return row_count

    @staticmethod
    def _require_width(columns: tuple[str, ...], row: tuple[Any, ...]) -> None:
        if len(row) != len(columns):
            raise DomainError(
                "ANALYTICS_EXPORT_CONTENT_WIDTH_INVALID",
                "the server export renderer received an invalid row width",
                409,
            )
