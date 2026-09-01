from __future__ import annotations

import hashlib
import os
import re
import subprocess
import time
from collections.abc import Callable
from pathlib import Path

from openpyxl import load_workbook

from .config import AgentConfig
from .errors import RunnerError
from .models import FT_JIEQUN_TOOL_CODE, ToolCapability, ToolRunResult

ProcessRunner = Callable[..., subprocess.CompletedProcess[str]]


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def ft_jiequn_capability(config: AgentConfig) -> ToolCapability:
    disabled_reason: str | None = None
    configured_sha = config.ft_package_sha256.lower()
    if not configured_sha:
        disabled_reason = "尚未配置已发布 FT 工具包的 SHA-256"
    elif not config.python_runtime.is_file():
        disabled_reason = "本机 Python 运行环境不可用"
    elif not config.ft_package.is_file():
        disabled_reason = "本机尚未安装已发布 FT 工具包"
    else:
        try:
            if file_sha256(config.ft_package) != configured_sha:
                disabled_reason = "FT 工具包与配置的发布 SHA-256 不一致"
        except OSError:
            disabled_reason = "本机无法读取已发布 FT 工具包"
    return ToolCapability(
        tool_code=FT_JIEQUN_TOOL_CODE,
        display_name="杰群 FT 原始目录低内存 PAT",
        test_stage="FT",
        factory_code="JIEQUN",
        analysis_type="QUICK_PAT",
        input_contract_version="JIEQUN_UNIFIED_CSV_DIRECTORY_V1",
        output_contract_version="FT_PAT_RESULT_V1",
        entrypoint="factories.jiequn.pat_cleaner.generate_raw_pat",
        allowed_suffixes=(".csv",),
        enabled=disabled_reason is None,
        disabled_reason=disabled_reason,
        package_sha256=configured_sha or None,
        timeout_seconds=config.run_timeout_seconds,
        max_output_bytes=config.max_output_bytes,
    )


def cp_capability_gate() -> ToolCapability:
    return ToolCapability(
        tool_code="CP_RAW_QUICK_PAT",
        display_name="CP 原始目录快速 PAT",
        test_stage="CP",
        factory_code="UNAPPROVED",
        analysis_type="QUICK_PAT",
        input_contract_version="NOT_APPROVED",
        output_contract_version="NOT_APPROVED",
        entrypoint="NOT_APPROVED",
        allowed_suffixes=(),
        enabled=False,
        disabled_reason="现有 CP 工具尚无已批准的原始目录 Quick PAT 入口和输出合同",
        package_sha256=None,
        timeout_seconds=None,
        max_output_bytes=None,
    )


class FtJiequnQuickPatRunner:
    """Invoke the approved function from the immutable FT PYZ package."""

    def __init__(
        self,
        config: AgentConfig,
        *,
        process_runner: ProcessRunner = subprocess.run,
    ) -> None:
        self._config = config
        self._process_runner = process_runner

    def run(
        self,
        *,
        tool: ToolCapability,
        source_path: Path,
        output_root: Path,
        expected_source_file_count: int,
    ) -> ToolRunResult:
        current = ft_jiequn_capability(self._config)
        if tool.tool_code != FT_JIEQUN_TOOL_CODE or not current.enabled:
            raise RunnerError(
                "LOCAL_TOOL_DISABLED",
                current.disabled_reason or "所选本机工具未获批准",
                409,
            )
        if current.package_sha256 != tool.package_sha256:
            raise RunnerError(
                "LOCAL_TOOL_RELEASE_CHANGED",
                "FT 工具发布版本在任务创建后发生变化，请重新预览",
                409,
            )
        # Validate immediately before execution; the PYZ must be immutable for the run.
        if file_sha256(self._config.ft_package) != tool.package_sha256:
            raise RunnerError(
                "LOCAL_TOOL_SHA_MISMATCH", "FT 工具包 SHA-256 校验失败", 409
            )
        output_root.mkdir(parents=True, exist_ok=False)
        environment = _isolated_environment(
            {
                "TMS_LOCAL_PAT_PACKAGE": str(self._config.ft_package),
                "TMS_LOCAL_PAT_INPUT": str(source_path),
                "TMS_LOCAL_PAT_OUTPUT": str(output_root),
                "PYTHONUTF8": "1",
                "PYTHONIOENCODING": "utf-8",
            }
        )
        started = time.perf_counter()
        try:
            completed = self._process_runner(
                [str(self._config.python_runtime), "-c", _JIEQUN_PAT_SCRIPT],
                # Any incidental relative output from the released tool stays inside
                # this run's isolated workspace rather than beside the source/PYZ.
                cwd=str(output_root),
                env=environment,
                text=True,
                encoding="utf-8",
                errors="replace",
                capture_output=True,
                check=False,
                timeout=self._config.run_timeout_seconds,
            )
        except subprocess.TimeoutExpired as exc:
            raise RunnerError(
                "LOCAL_TOOL_TIMEOUT", "本机 FT PAT 计算超时", 504
            ) from exc
        except OSError as exc:
            raise RunnerError(
                "LOCAL_TOOL_START_FAILED", "本机 FT PAT 工具无法启动", 500
            ) from exc
        elapsed = time.perf_counter() - started
        if completed.returncode != 0:
            raise RunnerError(
                "LOCAL_TOOL_FAILED",
                "既有 FT PAT 工具执行失败，请在本机检查源格式和 Agent 日志",
                422,
            )
        reports = sorted(
            output_root.rglob("PAT_*.xlsx"), key=lambda item: str(item).casefold()
        )
        if len(reports) != 1:
            raise RunnerError(
                "LOCAL_RESULT_CONTRACT_INVALID",
                "FT PAT 输出合同要求且只允许一个 PAT Excel",
                500,
            )
        report = reports[0].resolve()
        try:
            report.relative_to(output_root.resolve())
        except ValueError as exc:
            raise RunnerError(
                "LOCAL_RESULT_PATH_ESCAPE", "FT PAT 结果超出本机工作目录", 500
            ) from exc
        if report.stat().st_size > self._config.max_output_bytes:
            raise RunnerError(
                "LOCAL_RESULT_TOO_LARGE", "FT PAT 结果超过已批准大小上限", 507
            )
        parameter_count = _read_parameter_count(report)
        engine_files, record_count, engine_parameters = _parse_engine_summary(
            completed.stdout or ""
        )
        if (
            engine_files != expected_source_file_count
            or record_count < 1
            or engine_parameters != parameter_count
        ):
            raise RunnerError(
                "LOCAL_RESULT_COUNT_MISMATCH", "FT PAT 运行摘要与结果工作簿不一致", 500
            )
        return ToolRunResult(
            report_path=report,
            parameter_count=parameter_count,
            record_count=record_count,
            elapsed_seconds=round(elapsed, 3),
            stdout_tail=(completed.stdout or "")[-4000:],
        )


def _isolated_environment(extra: dict[str, str]) -> dict[str, str]:
    allowed = (
        "SystemRoot",
        "WINDIR",
        "COMSPEC",
        "TEMP",
        "TMP",
        "PATH",
        "PATHEXT",
        "NUMBER_OF_PROCESSORS",
    )
    environment = {key: os.environ[key] for key in allowed if key in os.environ}
    environment.update(extra)
    return environment


def _read_parameter_count(path: Path) -> int:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl exposes several parser exception types
        raise RunnerError(
            "LOCAL_RESULT_WORKBOOK_INVALID", "FT PAT 结果不是有效的 XLSX 工作簿", 500
        ) from exc
    try:
        sheet = workbook.active
        count = 0
        for index, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            if index <= 2:
                continue
            parameter = str(values[0]).strip() if values and values[0] is not None else ""
            if parameter:
                count += 1
    finally:
        workbook.close()
    if count < 1:
        raise RunnerError(
            "LOCAL_RESULT_WORKBOOK_EMPTY", "FT PAT 结果没有参数统计行", 500
        )
    return count


def _parse_engine_summary(stdout: str) -> tuple[int, int, int]:
    matches = re.findall(
        r"PAT [^\r\n]*?:\s*文件=([\d,]+),\s*解析行=([\d,]+),\s*参数=(\d+)",
        stdout,
    )
    if len(matches) != 1:
        raise RunnerError(
            "LOCAL_TOOL_SUMMARY_MISSING", "FT PAT 工具没有输出唯一可审计运行摘要", 500
        )
    files, rows, parameters = matches[0]
    return int(files.replace(",", "")), int(rows.replace(",", "")), int(parameters)


_JIEQUN_PAT_SCRIPT = """
import os, sys
from pathlib import Path
package = os.environ['TMS_LOCAL_PAT_PACKAGE']
source = Path(os.environ['TMS_LOCAL_PAT_INPUT']).resolve()
output = Path(os.environ['TMS_LOCAL_PAT_OUTPUT']).resolve()
sys.path.insert(0, package)
from factories.jiequn.dc_auto import DC_FORMAT_UNIFIED, detect_dc_format
from factories.jiequn.pat_cleaner import generate_raw_pat
detection = detect_dc_format(source)
if detection.format_name != DC_FORMAT_UNIFIED:
    raise SystemExit('Local Quick PAT only accepts the approved Jiequn unified CSV contract')
detected = {path.resolve() for path in detection.files}
all_csv = {
    path.resolve() for path in source.rglob('*')
    if path.is_file() and path.suffix.lower() == '.csv'
}
if detected != all_csv:
    raise SystemExit(
        f'Jiequn unified CSV contract mismatch: detected={len(detected)}, csv={len(all_csv)}'
    )
result = generate_raw_pat(source_dir=source, output_dir=output)
if not result:
    raise SystemExit('Jiequn Quick PAT returned no result')
print(f'TMS_LOCAL_PAT_RESULT={Path(result).resolve()}')
"""
