from __future__ import annotations

import csv
import hashlib
import io
import os
import time
from contextlib import AbstractContextManager, nullcontext
from dataclasses import replace
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from app.core.errors import DomainError
from app.domain.analytics import AnalyticsContextRequest
from app.domain.analytics_export_worker import (
    AnalyticsExportTable,
    AnalyticsExportWorkItem,
    RenderedAnalyticsExport,
)
from app.domain.analytics_exports import (
    AnalyticsExportDatasetRecord,
    AnalyticsExportFormat,
    AnalyticsExportRecord,
    AnalyticsExportScope,
)
from app.domain.analytics_risk import (
    AnalyticsEvaluatedRiskItem,
    AnalyticsRiskRuleProvenance,
)
from app.domain.saved_analyses import (
    SavedAnalysisRuleContext,
    canonical_json,
    saved_analysis_hashes,
    validate_analysis_presentation_config,
)
from app.infrastructure import analytics_export_renderer as renderer_module
from app.infrastructure import sql_analytics_export_content as content_module
from app.infrastructure.analytics_export_files import (
    AnalyticsExportPathPolicy,
    UnsafeAnalyticsExportPath,
)
from app.infrastructure.analytics_export_renderer import AnalyticsExportRenderer
from app.infrastructure.sql_analytics_export_content import (
    SqlAnalyticsExportContentSource,
)
from app.infrastructure.sql_analytics_export_service import SqlAnalyticsExportService
from app.infrastructure.sql_analytics_export_worker import (
    SqlAnalyticsExportWorkerRepository,
)
from app.workers.analytics_export_worker import AnalyticsExportWorker
from openpyxl import load_workbook
from PIL import Image as PillowImage


def _work_item(
    *,
    export_format: str = "CSV",
    export_scope: str = "FILTERED_RESULT",
    template_code: str = "ANALYTICS_DETAIL",
    export_job_id: int = 41,
    parameters: tuple[str, ...] = ("IDSS",),
) -> AnalyticsExportWorkItem:
    chart_config = {"show_spec_overlay": True, "y_axis_min": 0.0}
    report_analysis = {
        "ANALYTICS_OVERVIEW": {
            "contract_version": "ANALYTICS_EXPORT_ANALYSIS_CONFIG_V1",
            "section": "OVERVIEW",
            "overview": {"evaluations": []},
        },
        "PARAMETER_ANALYSIS": {
            "contract_version": "ANALYTICS_EXPORT_ANALYSIS_CONFIG_V1",
            "section": "PARAMETER_ANALYSIS",
            "parameter_analysis": {
                "parameters": list(parameters),
                "group_by": "DATASET",
                "analyses": ["DESCRIPTIVE"],
            },
        },
        "PARAMETER_RELATIONSHIP": {
            "contract_version": "ANALYTICS_EXPORT_ANALYSIS_CONFIG_V1",
            "section": "PARAMETER_RELATIONSHIP",
            "parameter_relationship": {
                "x_parameter": parameters[0] if parameters else "IDSS",
                "y_parameters": list(parameters[1:] or ("VTH",)),
                "analyses": ["SCATTER"],
                "group_by": "DATASET",
            },
        },
        "SPATIAL_ANALYSIS": {
            "contract_version": "ANALYTICS_EXPORT_ANALYSIS_CONFIG_V1",
            "section": "SPATIAL_ANALYSIS",
            "spatial_analysis": {"mode": "BIN_MAP"},
        },
        "FT_QUALITY": {
            "contract_version": "ANALYTICS_EXPORT_ANALYSIS_CONFIG_V1",
            "section": "FT_QUALITY",
            "ft_quality": {
                "analysis": "SYL_GROUPED_LIMIT",
                "rule": {"rule_code": "SYL_GROUPED_LIMIT", "version_code": "V1"},
                "group_by": "DATASET",
            },
        },
        "WAFER_SUMMARY": {
            "contract_version": "ANALYTICS_EXPORT_ANALYSIS_CONFIG_V1",
            "section": "WAFER_SUMMARY",
            "wafer_summary": {"sort_by": "DATASET", "sort_direction": "ASC"},
        },
    }.get(template_code)
    if export_scope == "REPORT" and report_analysis is not None:
        chart_config["analysis"] = report_analysis
    display_config = {
        "section": "detail",
        "page": 1,
        "page_size": 50,
        "focus_dataset_id": 7,
    }
    if export_scope == "CURRENT_PAGE":
        chart_config["analysis_view_state"] = {
            "contract_version": "ANALYSIS_VIEW_STATE_V1",
            "components": {
                "detail": {
                    "view": "WIDE",
                    "sortBy": "UNIT_SEQUENCE",
                    "sortDirection": "ASC",
                }
            },
        }
    return AnalyticsExportWorkItem(
        export_job_id=export_job_id,
        requested_by=10,
        export_scope=AnalyticsExportScope(export_scope),
        export_format=AnalyticsExportFormat(export_format),
        template_code=template_code,
        template_version="v1",
        context=AnalyticsContextRequest.model_validate(
            {
                "datasets": [{"dataset_id": 7, "version_no": 3}],
                "filters": {"lot_ids": ["LOT-1"]},
                "parameters": list(parameters),
            }
        ),
        dataset_version_ids=(71,),
        test_stage="CP",
        filter_hash="a" * 64,
        context_hash="b" * 64,
        rule_context=SavedAnalysisRuleContext(
            spec_versions=["SPEC:1:v1"],
            bin_mapping_versions=[],
            evaluation_rule_versions=[],
        ),
        chart_config=chart_config,
        display_config=display_config,
        presentation_hash=validate_analysis_presentation_config(
            chart_config, display_config
        ),
        artifact_ttl_hours=24,
        page=1 if export_scope == "CURRENT_PAGE" else None,
        page_size=50 if export_scope == "CURRENT_PAGE" else None,
        requested_at_utc=datetime.now(UTC),
        lease_token="11111111-1111-4111-8111-111111111111",
        lease_owner="worker-test",
        lease_expires_at_utc=datetime.now(UTC) + timedelta(minutes=5),
        attempt_count=1,
    )


class _Content:
    def table(self, _work_item) -> AnalyticsExportTable:
        return AnalyticsExportTable(
            columns=("lot_id", "value", "comment"),
            rows=iter(
                (
                    ("LOT-1", 1.25, "normal"),
                    ("LOT-2", None, "=unsafe-formula"),
                )
            ),
        )


@pytest.mark.parametrize(
    ("export_format", "scope", "template", "suffix"),
    (
        ("CSV", "FILTERED_RESULT", "ANALYTICS_DETAIL", ".csv"),
        ("XLSX", "FILTERED_RESULT", "ANALYTICS_DETAIL", ".xlsx"),
        ("BIN_TXT", "FILTERED_RESULT", "ANALYTICS_DETAIL", ".txt"),
        ("HTML", "REPORT", "ANALYTICS_OVERVIEW", ".html"),
        ("PDF", "REPORT", "ANALYTICS_OVERVIEW", ".pdf"),
        ("PNG", "REPORT", "ANALYTICS_OVERVIEW", ".png"),
    ),
)
def test_renderer_generates_supported_formats_from_one_content_contract(
    tmp_path: Path,
    export_format: str,
    scope: str,
    template: str,
    suffix: str,
) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    item = _work_item(
        export_format=export_format,
        export_scope=scope,
        template_code=template,
    )
    artifact = AnalyticsExportRenderer(policy, _Content()).render(item)

    assert artifact.path.suffix == suffix
    assert artifact.exported_row_count == 2
    assert artifact.file_size == artifact.path.stat().st_size
    assert artifact.sha256 == hashlib.sha256(artifact.path.read_bytes()).hexdigest()
    assert not list(artifact.path.parent.glob("*.tmp"))

    if export_format == "CSV":
        rows = list(
            csv.reader(io.StringIO(artifact.path.read_text(encoding="utf-8-sig")))
        )
        assert rows[0] == ["lot_id", "value", "comment"]
        assert rows[2][2] == "'=unsafe-formula"
    elif export_format == "XLSX":
        workbook = load_workbook(artifact.path, read_only=True, data_only=False)
        assert workbook.sheetnames == ["Context", "Data"]
        assert list(workbook["Data"].values)[2][2] == "'=unsafe-formula"
        assert dict(workbook["Context"].values)["context_hash"] == "b" * 64
        assert (
            dict(workbook["Context"].values)["presentation_hash"]
            == item.presentation_hash
        )
        workbook.close()
    elif export_format == "BIN_TXT":
        assert "lot_id\tvalue\tcomment" in artifact.path.read_text(encoding="utf-8")
    elif export_format == "HTML":
        rendered = artifact.path.read_text(encoding="utf-8")
        assert "TMS Analytics Export" in rendered
        assert "&#x27;=unsafe-formula" in rendered
        assert item.presentation_hash in rendered
    elif export_format == "PDF":
        rendered = artifact.path.read_bytes()
        assert rendered.startswith(b"%PDF-")
        assert rendered.rstrip().endswith(b"%%EOF")
    else:
        assert artifact.path.read_bytes().startswith(b"\x89PNG\r\n\x1a\n")
        with PillowImage.open(artifact.path) as image:
            assert image.format == "PNG"
            assert image.width >= 1_000
            image.verify()


def test_xlsx_context_splits_values_at_excel_cell_limit(tmp_path: Path) -> None:
    item = _work_item(export_format="XLSX")
    filters = item.context.filters.model_copy(
        update={
            "lot_ids": [f"LOT-{index:02d}-" + ("x" * 190) for index in range(50)],
            "wafer_ids": [f"WAFER-{index:03d}-" + ("x" * 187) for index in range(100)],
            "source_ids": [f"SOURCE-{index:02d}-" + ("x" * 187) for index in range(50)],
        }
    )
    context = item.context.model_copy(update={"filters": filters})
    item = replace(item, context=context)

    artifact = AnalyticsExportRenderer(
        AnalyticsExportPathPolicy(tmp_path / "exports"), _Content()
    ).render(item)
    workbook = load_workbook(artifact.path, read_only=True, data_only=False)
    context_rows = list(workbook["Context"].values)
    workbook.close()

    filter_chunks = [
        value
        for key, value in context_rows
        if isinstance(key, str) and key.startswith("filters")
    ]
    assert len(filter_chunks) >= 2
    assert all(
        isinstance(chunk, str) and len(chunk) <= 30_000 for chunk in filter_chunks
    )
    assert "".join(filter_chunks) == filters.model_dump_json(exclude_none=True)


@pytest.mark.parametrize("export_format", ("CSV", "XLSX", "BIN_TXT"))
def test_tabular_renderer_neutralizes_formula_like_headers(
    tmp_path: Path, export_format: str
) -> None:
    class _UnsafeHeaderContent:
        @staticmethod
        def table(_work_item) -> AnalyticsExportTable:
            return AnalyticsExportTable(
                columns=("=unsafe-header",), rows=iter((("\t=unsafe-value",),))
            )

    artifact = AnalyticsExportRenderer(
        AnalyticsExportPathPolicy(tmp_path / "exports"), _UnsafeHeaderContent()
    ).render(_work_item(export_format=export_format))

    if export_format == "XLSX":
        workbook = load_workbook(artifact.path, read_only=True, data_only=False)
        values = list(workbook["Data"].values)
        workbook.close()
        assert values == [("'=unsafe-header",), ("'\t=unsafe-value",)]
    else:
        delimiter = "," if export_format == "CSV" else "\t"
        encoding = "utf-8-sig" if export_format == "CSV" else "utf-8"
        with artifact.path.open("r", encoding=encoding, newline="") as stream:
            values = list(csv.reader(stream, delimiter=delimiter))
        assert values == [["'=unsafe-header"], ["'\t=unsafe-value"]]


@pytest.mark.parametrize(
    ("export_format", "sentinel"), (("PDF", "pdf_canvas"), ("PNG", "Image"))
)
def test_missing_pdf_and_png_dependencies_fail_stably(
    tmp_path: Path, monkeypatch, export_format: str, sentinel: str
) -> None:
    monkeypatch.setattr(renderer_module, sentinel, None)
    renderer = AnalyticsExportRenderer(
        AnalyticsExportPathPolicy(tmp_path / "exports"), _Content()
    )
    with pytest.raises(DomainError) as failure:
        renderer.render(
            _work_item(
                export_format=export_format,
                export_scope="REPORT",
                template_code="ANALYTICS_OVERVIEW",
            )
        )
    assert failure.value.code == "ANALYTICS_EXPORT_DEPENDENCY_UNAVAILABLE"
    assert not (tmp_path / "exports").exists()


@pytest.mark.parametrize("export_format", ("PDF", "PNG"))
def test_pdf_and_png_use_deterministic_ascii_fallback_without_system_font(
    tmp_path: Path, monkeypatch, export_format: str
) -> None:
    class _ChineseContent:
        @staticmethod
        def table(_work_item) -> AnalyticsExportTable:
            return AnalyticsExportTable(columns=("指标",), rows=iter((("良率",),)))

    monkeypatch.setattr(renderer_module, "_font_candidates", lambda: ())
    artifact = AnalyticsExportRenderer(
        AnalyticsExportPathPolicy(tmp_path / "exports"), _ChineseContent()
    ).render(
        _work_item(
            export_format=export_format,
            export_scope="REPORT",
            template_code="ANALYTICS_OVERVIEW",
        )
    )

    assert artifact.exported_row_count == 1
    assert renderer_module._ascii_fallback("指标", unicode_font=False) == (
        r"\u6307\u6807"
    )
    assert artifact.file_size > 0


@pytest.mark.parametrize("export_format", ("PDF", "PNG"))
def test_report_media_counts_all_rows_while_bounding_preview(
    tmp_path: Path, export_format: str
) -> None:
    class _ManyRows:
        @staticmethod
        def table(_work_item) -> AnalyticsExportTable:
            return AnalyticsExportTable(
                columns=("metric", "count"),
                rows=(("ROW", index) for index in range(75)),
            )

    artifact = AnalyticsExportRenderer(
        AnalyticsExportPathPolicy(tmp_path / "exports"), _ManyRows()
    ).render(
        _work_item(
            export_format=export_format,
            export_scope="REPORT",
            template_code="ANALYTICS_OVERVIEW",
        )
    )

    assert artifact.exported_row_count == 75
    assert artifact.file_size > 0


def test_atomic_renderer_removes_partial_file_after_content_failure(
    tmp_path: Path,
) -> None:
    class _FailingContent:
        @staticmethod
        def table(_work_item) -> AnalyticsExportTable:
            def rows():
                yield ("LOT-1",)
                raise RuntimeError("synthetic content failure")

            return AnalyticsExportTable(columns=("lot_id",), rows=rows())

    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    with pytest.raises(RuntimeError, match="synthetic content failure"):
        AnalyticsExportRenderer(policy, _FailingContent()).render(_work_item())

    assert not policy.job_root(41).exists()


class _ReportRoutingContent(SqlAnalyticsExportContentSource):
    def __init__(self) -> None:
        super().__init__(None)  # type: ignore[arg-type]

    def _iter_analysis_report_rows(self, work_item, config):
        yield self._analysis_row(
            "ROUTED",
            {"section": config.section.value},
            dataset_id=7,
            version_no=3,
            metric_value=work_item.template_code,
        )


@pytest.mark.parametrize(
    ("template", "parameters"),
    (
        ("ANALYTICS_OVERVIEW", ("IDSS",)),
        ("FT_QUALITY", ("IDSS",)),
        ("WAFER_SUMMARY", ("IDSS",)),
        ("PARAMETER_ANALYSIS", ("IDSS",)),
        ("PARAMETER_RELATIONSHIP", ("IDSS", "VTH")),
        ("SPATIAL_ANALYSIS", ()),
    ),
)
def test_every_registered_report_template_has_server_content(
    template: str, parameters: tuple[str, ...]
) -> None:
    table = _ReportRoutingContent().table(
        _work_item(
            export_format="CSV",
            export_scope="REPORT",
            template_code=template,
            parameters=parameters,
        )
    )
    assert table.columns[0] == "record_type"
    assert len(next(table.rows)) == len(table.columns)


def test_relationship_report_requires_exact_analysis_config() -> None:
    item = _work_item(
        export_format="CSV",
        export_scope="REPORT",
        template_code="PARAMETER_RELATIONSHIP",
    )
    item.chart_config.pop("analysis")
    with pytest.raises(DomainError) as failure:
        _ReportRoutingContent().table(item)
    assert failure.value.code == "ANALYTICS_EXPORT_ANALYSIS_CONFIG_INVALID"


@pytest.mark.parametrize(
    ("template", "parameters", "service_name", "method_name", "result"),
    (
        (
            "ANALYTICS_OVERVIEW",
            ("IDSS",),
            "SqlAnalyticsService",
            "overview",
            {
                "datasets": (),
                "yield_trend": (),
                "bin_pareto": (),
                "wafer_map": (),
                "risk_summary": (),
            },
        ),
        (
            "PARAMETER_ANALYSIS",
            ("IDSS",),
            "SqlDatasetService",
            "analyze_parameters",
            {"group_by": "DATASET", "items": ()},
        ),
        (
            "PARAMETER_RELATIONSHIP",
            ("IDSS", "VTH"),
            "SqlParameterRelationshipService",
            "relationship",
            {"group_by": "DATASET", "trend_order_basis": "RUN", "items": ()},
        ),
        (
            "SPATIAL_ANALYSIS",
            (),
            "SqlSpatialAnalysisService",
            "analyze",
            {
                "mode": "BIN_MAP",
                "parameter": None,
                "data_quality": {},
                "color_domain": None,
                "zone_geometry": None,
                "wafer_manifest": (),
                "points": (),
                "wafer_layers": (),
                "zones": (),
                "quadrants": (),
            },
        ),
        (
            "FT_QUALITY",
            ("IDSS",),
            "SqlQualityEvaluationService",
            "analyze",
            {
                "analysis": "SYL_GROUPED_LIMIT",
                "calculation_context_hash": "c" * 64,
                "rule": {},
                "parameter_identity": None,
                "pat": (),
                "spc": (),
                "margin": (),
                "bin_cooccurrence": (),
                "sbl": (),
                "syl": (),
                "pass_fail_distribution": (),
            },
        ),
        (
            "WAFER_SUMMARY",
            ("IDSS",),
            "SqlWaferSummaryService",
            "summarize",
            {
                "sort_by": "DATASET",
                "sort_direction": "ASC",
                "items": (),
                "total": 0,
            },
        ),
    ),
)
def test_each_report_template_invokes_its_authoritative_analysis_service(
    monkeypatch,
    template: str,
    parameters: tuple[str, ...],
    service_name: str,
    method_name: str,
    result: dict,
) -> None:
    from types import SimpleNamespace

    requests = []

    class _Service:
        def __init__(self, *_args, **_kwargs) -> None:
            pass

    def execute(_self, request):
        requests.append(request)
        return SimpleNamespace(**result)

    setattr(_Service, method_name, execute)
    monkeypatch.setattr(content_module, service_name, _Service)
    source = SqlAnalyticsExportContentSource(None)  # type: ignore[arg-type]
    context_row = source._analysis_row("TEST_CONTEXT", {"authoritative": True})
    monkeypatch.setattr(
        source,
        "_result_context_rows",
        lambda *_args, **_kwargs: iter((context_row,)),
    )
    table = source.table(
        _work_item(
            export_format="CSV",
            export_scope="REPORT",
            template_code=template,
            parameters=parameters,
        )
    )
    rows = list(table.rows)

    assert len(requests) == 1
    assert rows[0][0] == "TEST_CONTEXT"
    assert len(rows[0]) == len(table.columns)


def test_report_rule_context_mismatch_fails_closed() -> None:
    from types import SimpleNamespace

    item = _work_item(
        export_format="CSV",
        export_scope="REPORT",
        template_code="ANALYTICS_OVERVIEW",
    )
    result = SimpleNamespace(
        rule_context=SimpleNamespace(
            spec_versions=("SPEC:99:V2",),
            bin_mapping_versions=(),
            evaluation_rule_versions=(),
        )
    )

    with pytest.raises(DomainError) as failure:
        SqlAnalyticsExportContentSource._assert_rule_context(item, result)

    assert failure.value.code == "ANALYTICS_EXPORT_RULE_CONTEXT_STALE"


def test_quality_rule_gate_error_is_preserved_by_report_content(monkeypatch) -> None:
    class _QualityService:
        def __init__(self, *_args, **_kwargs) -> None:
            pass

        def analyze(self, _request):
            raise DomainError(
                "ANALYSIS_RULE_NOT_APPROVED",
                "requested quality rule is not approved",
                409,
            )

    monkeypatch.setattr(content_module, "SqlQualityEvaluationService", _QualityService)
    table = SqlAnalyticsExportContentSource(None).table(  # type: ignore[arg-type]
        _work_item(
            export_format="CSV",
            export_scope="REPORT",
            template_code="FT_QUALITY",
        )
    )

    with pytest.raises(DomainError) as failure:
        next(table.rows)

    assert failure.value.code == "ANALYSIS_RULE_NOT_APPROVED"


def test_overview_report_recomputes_pinned_instant_risk_without_client_results(
    monkeypatch,
) -> None:
    from types import SimpleNamespace

    base_item = _work_item(
        export_format="CSV",
        export_scope="REPORT",
        template_code="ANALYTICS_OVERVIEW",
    )
    chart_config = {
        **base_item.chart_config,
        "analysis": {
            "contract_version": "ANALYTICS_EXPORT_ANALYSIS_CONFIG_V1",
            "section": "OVERVIEW",
            "overview": {
                "evaluations": [
                    {
                        "analysis": "PAT_ROBUST_IQR",
                        "parameter": "IDSS",
                        "group_by": "DATASET",
                        "rule": {"rule_code": "CP_PAT", "version_code": "V1"},
                    }
                ]
            },
        },
    }
    item = replace(
        base_item,
        chart_config=chart_config,
        presentation_hash=validate_analysis_presentation_config(
            chart_config, base_item.display_config
        ),
        rule_context=SavedAnalysisRuleContext(
            spec_versions=["SPEC:1:v1"],
            evaluation_rule_versions=["RULE:CP_PAT:V1"],
        ),
    )
    overview_result = SimpleNamespace(
        datasets=(),
        yield_trend=(),
        bin_pareto=(),
        wafer_map=(),
        risk_summary=(),
    )

    class _OverviewService:
        def __init__(self, *_args, **_kwargs) -> None:
            pass

        def overview(self, _request):
            return overview_result

    risk_requests = []
    risk = AnalyticsEvaluatedRiskItem(
        "PAT:7:ABC",
        "PAT_ROBUST_IQR",
        "QUALITY",
        "WARNING",
        "ACTIVE",
        None,
        "PAT",
        "approved PAT result",
        7,
        3,
        "DATASET:7:V3",
        "IDSS",
        "PAT_OUTLIER_RATE",
        0.1,
        "> 0",
        0.0,
        1,
        10,
        0.1,
        ("UNIT:1",),
        False,
        AnalyticsRiskRuleProvenance(
            "CP_PAT", "V1", "PAT_SHARED_IQR_1_35_V1", "APPROVED", "ENABLED", "f" * 64
        ),
    )

    class _RiskService:
        def __init__(self, *_args, **_kwargs) -> None:
            pass

        def evaluate(self, request):
            risk_requests.append(request)
            return SimpleNamespace(
                contract_version="ANALYTICS_INSTANT_RISK_V1",
                filter_summary=SimpleNamespace(
                    filter_hash=item.filter_hash, context_hash=item.context_hash
                ),
                calculation_context_hash="c" * 64,
                requested_analyses=("PAT_ROBUST_IQR",),
                items=(risk,),
                warnings=(),
                computed_at="2026-08-31T00:00:00+00:00",
            )

    monkeypatch.setattr(content_module, "SqlAnalyticsService", _OverviewService)
    monkeypatch.setattr(content_module, "AnalyticsInstantRiskService", _RiskService)
    source = SqlAnalyticsExportContentSource(None)  # type: ignore[arg-type]
    monkeypatch.setattr(
        source,
        "_result_context_rows",
        lambda *_args, **_kwargs: iter(
            (source._analysis_row("TEST_CONTEXT", {"authoritative": True}),)
        ),
    )

    rows = list(source.table(item).rows)

    assert len(risk_requests) == 1
    assert risk_requests[0].evaluations[0].rule.rule_code == "CP_PAT"
    assert {row[0] for row in rows} >= {
        "INSTANT_RISK_CONTEXT",
        "INSTANT_RISK",
        "INSTANT_RISK_EVIDENCE",
    }


def test_full_dataset_scope_does_not_apply_screen_filters() -> None:
    item = _work_item(export_scope="FULL_DATASET")

    clause, parameters, expanding = SqlAnalyticsExportContentSource._filter_sql(
        item,
        source_run_ids=(101,),
        condition_item_ids=(201,),
    )

    assert clause == ""
    assert parameters == {}
    assert expanding == ()


class _StreamingPageContent(SqlAnalyticsExportContentSource):
    def __init__(self) -> None:
        super().__init__(None)  # type: ignore[arg-type]
        self.seen = 0
        self.detail_requests = []

    def _unit_stream(self, *_args, **kwargs):
        self.detail_requests.append((_args[2], kwargs.get("detail_request")))
        for unit_id in range(1, 101):
            self.seen += 1
            yield {
                "dataset_id": 7,
                "version_no": 3,
                "test_stage": "CP",
                "unit_id": unit_id,
                "logical_unit_key": f"UNIT-{unit_id}",
                "lot_id": "LOT-1",
                "wafer_id": "W1",
                "x_coord": unit_id,
                "y_coord": 1,
                "soft_bin": "1",
                "hard_bin": None,
                "overall_result": "PASS",
                "source_row_no": unit_id,
                "run_id": 1,
                "metadata_json": None,
                "tester_id": "T1",
                "program_version": "P1",
            }

    @staticmethod
    def _dataset_item_rows(_connection, _dataset_version_id):
        return (
            {
                "test_item_id": 501,
                "raw_item_name": "IDSS",
                "canonical_parameter_code": "IDSS",
                "step_code": "FT",
                "sequence_no": 1,
                "unit_code": "uA",
                "program_lsl": None,
                "program_usl": None,
                "condition_json": None,
            },
        )


class _PageEngine:
    @staticmethod
    def connect():
        return nullcontext(object())


def test_current_page_stream_stops_after_requested_units() -> None:
    source = _StreamingPageContent()
    source._engine = _PageEngine()  # type: ignore[assignment]
    item = _work_item(export_scope="CURRENT_PAGE")
    item = replace(
        item,
        context=item.context.model_copy(update={"parameters": []}),
        page=1,
        page_size=2,
        display_config={**item.display_config, "page": 1, "page_size": 2},
    )

    rows = list(source._iter_detail_rows(item))

    assert [row[3] for row in rows] == [1, 2]
    assert source.seen == 3
    assert source.detail_requests[0][0] == 71
    assert source.detail_requests[0][1].sort_by == "UNIT_SEQUENCE"


def test_current_page_applies_frozen_ui_sort_and_focus_dataset() -> None:
    source = _StreamingPageContent()
    source._engine = _PageEngine()  # type: ignore[assignment]
    item = _work_item(export_scope="CURRENT_PAGE")
    item = replace(
        item,
        context=AnalyticsContextRequest.model_validate(
            {
                **item.context.model_dump(mode="json"),
                "datasets": [
                    {"dataset_id": 7, "version_no": 3},
                    {"dataset_id": 8, "version_no": 4},
                ],
                "parameters": [],
            }
        ),
        dataset_version_ids=(71, 84),
        chart_config={
            **item.chart_config,
            "analysis_view_state": {
                "contract_version": "ANALYSIS_VIEW_STATE_V1",
                "components": {
                    "detail": {
                        "view": "LONG",
                        "sortBy": "RESULT",
                        "sortDirection": "DESC",
                    }
                },
            },
        },
        display_config={
            **item.display_config,
            "focus_dataset_id": 8,
            "page": 1,
            "page_size": 2,
        },
        page=1,
        page_size=2,
    )

    rows = list(source._iter_detail_rows(item))

    assert [row[3] for row in rows] == [1, 2]
    assert [call[0] for call in source.detail_requests] == [84]
    request = source.detail_requests[0][1]
    assert request.focus_dataset_id == 8
    assert request.view == "LONG"
    assert request.sort_by == "RESULT"
    assert request.sort_direction == "DESC"


def test_current_page_replays_evaluation_and_measurement_filters_into_sql() -> None:
    source = _StreamingPageContent()
    source._engine = _PageEngine()  # type: ignore[assignment]
    item = _work_item(export_scope="CURRENT_PAGE")
    detail_state = {
        "view": "LONG",
        "sortBy": "RESULT",
        "sortDirection": "DESC",
        "evaluation_filter": {
            "evaluation_type": "PAT",
            "evaluation_results": ["FAIL"],
            "rule_code": "PAT_ROBUST_IQR",
            "rule_version": "V2",
        },
        "measurement_filter": {
            "parameter": "IDSS",
            "lower_bound": -1.0,
            "upper_bound": 1.0,
            "lower_inclusive": False,
            "upper_inclusive": True,
        },
    }
    item = replace(
        item,
        chart_config={
            **item.chart_config,
            "analysis_view_state": {
                "contract_version": "ANALYSIS_VIEW_STATE_V1",
                "components": {"detail": detail_state},
            },
        },
        page=1,
        page_size=2,
        display_config={**item.display_config, "page": 1, "page_size": 2},
    )

    list(source._iter_detail_rows(item))
    request = source.detail_requests[0][1]
    assert request.evaluation_filter.evaluation_type == "PAT"
    assert request.evaluation_filter.evaluation_results == ["FAIL"]
    assert request.measurement_filter.parameter == "IDSS"
    assert request.measurement_filter.lower_inclusive is False

    item_rows = (
        {
            "test_item_id": 501,
            "raw_item_name": "IDSS",
            "canonical_parameter_code": "IDSS",
            "step_code": "FT",
            "sequence_no": 1,
            "unit_code": "uA",
            "program_lsl": None,
            "program_usl": None,
            "condition_json": None,
        },
    )
    sql, parameters, expanding = source._detail_scope_filter_sql(request, item_rows)

    assert "risk_me.evaluation_type=:detail_evaluation_type" in sql
    assert "risk_me.evaluation_result IN :detail_evaluation_results" in sql
    assert "risk_rs.rule_code=:detail_rule_code" in sql
    assert "aggregate_m.value_numeric>:aggregate_lower_bound" in sql
    assert "aggregate_m.value_numeric<=:aggregate_upper_bound" in sql
    assert parameters == {
        "detail_evaluation_type": "PAT",
        "detail_evaluation_results": ("FAIL",),
        "detail_rule_code": "PAT_ROBUST_IQR",
        "detail_rule_version": "V2",
        "detail_parameter_ids": (501,),
        "aggregate_parameter_ids": (501,),
        "aggregate_lower_bound": -1.0,
        "aggregate_upper_bound": 1.0,
    }
    assert expanding == (
        "detail_evaluation_results",
        "detail_parameter_ids",
        "aggregate_parameter_ids",
    )


def test_path_policy_blocks_escape(tmp_path: Path) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    policy.prepare_job_root(9)
    outside = tmp_path / "outside.csv"
    outside.write_text("outside", encoding="utf-8")

    with pytest.raises(UnsafeAnalyticsExportPath):
        policy.require_artifact(9, outside, must_exist=True)
    with pytest.raises(UnsafeAnalyticsExportPath):
        policy.artifact_path(9, "../escape.csv")


def test_path_policy_blocks_symlink(tmp_path: Path) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    root = policy.prepare_job_root(9)
    outside = tmp_path / "outside.csv"
    outside.write_text("outside", encoding="utf-8")
    linked = root / "linked.csv"
    try:
        os.symlink(outside, linked)
    except (OSError, NotImplementedError):
        pytest.skip("the current Windows account cannot create a test symlink")
    with pytest.raises(UnsafeAnalyticsExportPath):
        policy.require_artifact(9, linked, must_exist=True)


class _Repository:
    def __init__(self, item: AnalyticsExportWorkItem) -> None:
        self.item = item
        self.completed: RenderedAnalyticsExport | None = None
        self.failure: tuple[str, str] | None = None
        self.heartbeats = 0
        self.execution_checks = 0

    def claim_next(self):
        item, self.item = self.item, None
        return item

    def complete(self, _item, artifact, *, expires_at_utc):
        assert expires_at_utc > datetime.now(UTC)
        self.completed = artifact

    def assert_execution_authorized(self, _item):
        self.execution_checks += 1

    def heartbeat(self, _item):
        self.heartbeats += 1

    def fail(self, _item, *, error_code, error_message):
        self.failure = (error_code, error_message)


def test_worker_marks_supported_output_complete_and_dependency_failure_failed(
    tmp_path: Path, monkeypatch
) -> None:
    success_repository = _Repository(_work_item())
    worker = AnalyticsExportWorker(
        success_repository,
        AnalyticsExportRenderer(
            AnalyticsExportPathPolicy(tmp_path / "success"), _Content()
        ),
    )
    worker.run_once()
    assert success_repository.completed is not None
    assert success_repository.execution_checks == 1
    assert success_repository.failure is None

    monkeypatch.setattr(renderer_module, "pdf_canvas", None)
    failed_repository = _Repository(
        _work_item(
            export_format="PDF",
            export_scope="REPORT",
            template_code="ANALYTICS_OVERVIEW",
            export_job_id=42,
        )
    )
    worker = AnalyticsExportWorker(
        failed_repository,
        AnalyticsExportRenderer(
            AnalyticsExportPathPolicy(tmp_path / "failed"), _Content()
        ),
    )
    worker.run_once()
    assert failed_repository.completed is None
    assert failed_repository.failure is not None
    assert failed_repository.failure[0] == "ANALYTICS_EXPORT_DEPENDENCY_UNAVAILABLE"


def test_worker_stops_before_render_when_requester_is_disabled_after_claim(
    tmp_path: Path,
) -> None:
    class _DisabledRepository(_Repository):
        def assert_execution_authorized(self, _item):
            self.execution_checks += 1
            raise DomainError(
                "ANALYTICS_EXPORT_ACCESS_REVOKED",
                "requester account is disabled",
                409,
            )

    class _ForbiddenContent:
        @staticmethod
        def table(_item):
            raise AssertionError("disabled requester must not reach content rendering")

    repository = _DisabledRepository(_work_item(export_job_id=43))
    root = tmp_path / "disabled-before-render"
    AnalyticsExportWorker(
        repository,
        AnalyticsExportRenderer(AnalyticsExportPathPolicy(root), _ForbiddenContent()),
    ).run_once()

    assert repository.execution_checks == 1
    assert repository.failure is not None
    assert repository.failure[0] == "ANALYTICS_EXPORT_ACCESS_REVOKED"
    assert not (root / "43").exists()


def test_worker_heartbeats_during_render_and_discards_output_after_claim_loss(
    tmp_path: Path,
) -> None:
    class _SlowContent:
        @staticmethod
        def table(_item):
            def rows():
                time.sleep(0.05)
                yield ("LOT-1",)

            return AnalyticsExportTable(columns=("lot_id",), rows=rows())

    heartbeat_repository = _Repository(_work_item())
    AnalyticsExportWorker(
        heartbeat_repository,
        AnalyticsExportRenderer(
            AnalyticsExportPathPolicy(tmp_path / "heartbeat"), _SlowContent()
        ),
        heartbeat_seconds=0.01,
    ).run_once()
    assert heartbeat_repository.heartbeats >= 1
    assert heartbeat_repository.completed is not None

    class _LostRepository(_Repository):
        def complete(self, _item, artifact, *, expires_at_utc):
            raise DomainError(
                "ANALYTICS_EXPORT_WORKER_CLAIM_LOST",
                "synthetic fencing loss",
                409,
            )

    lost_repository = _LostRepository(_work_item(export_job_id=99))
    root = tmp_path / "lost"
    AnalyticsExportWorker(
        lost_repository,
        AnalyticsExportRenderer(AnalyticsExportPathPolicy(root), _Content()),
    ).run_once()
    assert lost_repository.completed is None
    assert lost_repository.failure is None
    assert not (root / "99").exists()


def test_worker_discards_rendered_output_and_marks_failed_when_access_is_revoked(
    tmp_path: Path,
) -> None:
    class _RevokedRepository(_Repository):
        def complete(self, _item, _artifact, *, expires_at_utc):
            assert expires_at_utc > datetime.now(UTC)
            raise DomainError(
                "ANALYTICS_EXPORT_ACCESS_REVOKED",
                "requester grant was revoked while rendering",
                409,
            )

    repository = _RevokedRepository(_work_item(export_job_id=100))
    root = tmp_path / "revoked"

    AnalyticsExportWorker(
        repository,
        AnalyticsExportRenderer(AnalyticsExportPathPolicy(root), _Content()),
    ).run_once()

    assert repository.completed is None
    assert repository.failure is not None
    assert repository.failure[0] == "ANALYTICS_EXPORT_ACCESS_REVOKED"
    assert not (root / "100").exists()


def test_cleanup_io_failure_keeps_job_failed_and_emits_controlled_log(
    tmp_path: Path,
    caplog,
) -> None:
    class _RevokedRepository(_Repository):
        def complete(self, _item, _artifact, *, expires_at_utc):
            assert expires_at_utc > datetime.now(UTC)
            raise DomainError(
                "ANALYTICS_EXPORT_ACCESS_REVOKED",
                "requester was disabled while rendering",
                409,
            )

    class _CleanupFailureRenderer(AnalyticsExportRenderer):
        @staticmethod
        def discard_attempt(_work_item):
            raise OSError("synthetic managed-directory cleanup failure")

    repository = _RevokedRepository(_work_item(export_job_id=101))
    root = tmp_path / "cleanup-failure"
    caplog.set_level("ERROR", logger="app.workers.analytics_export_worker")

    AnalyticsExportWorker(
        repository,
        _CleanupFailureRenderer(AnalyticsExportPathPolicy(root), _Content()),
    ).run_once()

    assert repository.failure is not None
    assert repository.failure[0] == "ANALYTICS_EXPORT_ACCESS_REVOKED"
    assert (root / "101").exists()
    assert any(
        record.message == "failed to clean rejected Analytics Export attempt"
        and record.export_job_id == 101
        and record.attempt_count == 1
        for record in caplog.records
    )


def test_attempt_specific_artifacts_prevent_stale_worker_file_collision(
    tmp_path: Path,
) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    renderer = AnalyticsExportRenderer(policy, _Content())
    first = renderer.render(_work_item(export_job_id=77))
    second = renderer.render(replace(_work_item(export_job_id=77), attempt_count=2))

    assert first.path != second.path
    assert first.path.is_file() and second.path.is_file()
    assert {first.path.name, second.path.name} == {
        "analytics-export-77-attempt-1.csv",
        "analytics-export-77-attempt-2.csv",
    }


def test_attempt_cleanup_removes_half_products_but_preserves_newer_attempt(
    tmp_path: Path,
) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    renderer = AnalyticsExportRenderer(policy, _Content())
    root = policy.prepare_job_root(78)
    stale_target = root / "analytics-export-78-attempt-1.csv"
    stale_temp = root / (
        ".analytics-export-78-attempt-1.csv."
        "0123456789abcdef0123456789abcdef.tmp"
    )
    newer_target = root / "analytics-export-78-attempt-2.csv"
    stale_target.write_bytes(b"stale")
    stale_temp.write_bytes(b"partial")
    newer_target.write_bytes(b"newer")

    renderer.discard_attempt(_work_item(export_job_id=78))

    assert not stale_target.exists()
    assert not stale_temp.exists()
    assert newer_target.read_bytes() == b"newer"


class _ScalarResult:
    def __init__(self, value=None) -> None:
        self.value = value

    def scalar_one_or_none(self):
        return self.value

    def scalar_one(self):
        return self.value


class _OptionalMappingResult:
    def __init__(self, row=None) -> None:
        self.row = row

    def mappings(self):
        return self

    def one_or_none(self):
        return self.row


class _RowcountResult:
    def __init__(self, rowcount: int) -> None:
        self.rowcount = rowcount


class _ClaimState:
    def __init__(self) -> None:
        self.available = True
        self.sql: list[str] = []


class _StoredWorkItemConnection:
    def __init__(
        self,
        *,
        tamper_presentation: bool = False,
        current_page: bool = False,
        include_frozen_detail: bool = False,
        tamper_frozen_detail: bool = False,
        authorized: bool = True,
    ) -> None:
        context = AnalyticsContextRequest.model_validate(
            {
                "datasets": [{"dataset_id": 7, "version_no": 3}],
                "filters": {"lot_ids": ["LOT-1"]},
                "parameters": ["IDSS"],
            }
        )
        hashes = saved_analysis_hashes(context)
        chart_config = {"show_spec_overlay": True, "correlation_min_abs": 0.4}
        display_config = {"section": "parameter", "page": 1, "page_size": 50}
        detail_state = None
        if current_page:
            detail_state = {
                "view": "LONG",
                "sortBy": "RESULT",
                "sortDirection": "DESC",
                "evaluation_filter": {
                    "evaluation_type": "PAT",
                    "evaluation_results": ["FAIL"],
                    "rule_code": "PAT_ROBUST_IQR",
                    "rule_version": "V2",
                },
                "measurement_filter": {
                    "parameter": "IDSS",
                    "lower_bound": -1.0,
                    "upper_bound": 1.0,
                    "lower_inclusive": True,
                    "upper_inclusive": False,
                },
            }
            chart_config["analysis_view_state"] = {
                "contract_version": "ANALYSIS_VIEW_STATE_V1",
                "components": {"detail": detail_state},
            }
            display_config = {
                "section": "detail",
                "page": 1,
                "page_size": 50,
                "focus_dataset_id": 7,
            }
        presentation_hash = validate_analysis_presentation_config(
            chart_config, display_config
        )
        if tamper_presentation:
            chart_config["correlation_min_abs"] = 0.9
        envelope = {
            "artifact_ttl_hours": 24,
            "chart_config": chart_config,
            "display_config": display_config,
            "filters": hashes.normalized_filters,
            "page": 1 if current_page else None,
            "page_size": 50 if current_page else None,
            "parameters": list(hashes.normalized_parameters),
            "presentation_hash": presentation_hash,
            "request_reason_sha256": "c" * 64,
        }
        if include_frozen_detail:
            assert detail_state is not None
            envelope["current_page_detail_state"] = {
                **detail_state,
                **({"view": "WIDE"} if tamper_frozen_detail else {}),
            }
        self.job = {
            "export_job_id": 41,
            "requested_by": 10,
            "dataset_version_id": 71,
            "export_scope": "CURRENT_PAGE" if current_page else "FILTERED_RESULT",
            "export_format": "CSV",
            "template_code": "ANALYTICS_DETAIL",
            "template_version": "v1",
            "filter_json": canonical_json(envelope),
            "status": "RUNNING",
            "requested_at_utc": datetime.now(UTC),
            "contract_version": "ANALYTICS_EXPORT_V1",
            "filter_hash": hashes.filter_hash,
            "context_hash": hashes.context_hash,
            "rule_context_json": SavedAnalysisRuleContext().model_dump_json(),
            "lease_token": "11111111-1111-4111-8111-111111111111",
            "lease_owner": "worker-test",
            "lease_expires_at_utc": datetime.now(UTC) + timedelta(minutes=5),
            "attempt_count": 1,
        }
        self.dataset = {
            "dataset_version_id": 71,
            "ordinal_no": 1,
            "dataset_id": 7,
            "version_no": 3,
            "status": "PUBLISHED",
            "is_current": True,
            "test_stage": "CP",
            "requested_by_user_id": 10,
            "job_status": "RUNNING",
            "job_contract_version": "ANALYTICS_EXPORT_V1",
            "job_lease_token": "11111111-1111-4111-8111-111111111111",
            "job_lease_owner": "worker-test",
            "job_lease_expires_at_utc": datetime.now(UTC)
            + timedelta(minutes=5),
            "can_read": authorized,
        }
        self.calls: list[str] = []

    def execute(self, statement, _parameters=None):
        sql = " ".join(str(statement).split())
        self.calls.append(sql)
        if sql.startswith("SELECT export_job_id,requested_by,dataset_version_id"):
            return _OptionalMappingResult(self.job)
        if sql.startswith("SELECT ejd.dataset_version_id"):
            return _MappingRowsResult([self.dataset])
        raise AssertionError(sql)


class _MappingRowsResult:
    def __init__(self, rows) -> None:
        self.rows = rows

    def mappings(self):
        return self

    def all(self):
        return self.rows


def test_sql_worker_round_trips_and_verifies_presentation_envelope(
    tmp_path: Path,
) -> None:
    repository = SqlAnalyticsExportWorkerRepository(
        object(),  # type: ignore[arg-type]
        AnalyticsExportPathPolicy(tmp_path / "exports"),
        worker_id="worker-test",
    )
    item = repository._load_work_item(
        _StoredWorkItemConnection(),  # type: ignore[arg-type]
        41,
        expected_lease_token="11111111-1111-4111-8111-111111111111",
    )

    assert item.chart_config["correlation_min_abs"] == 0.4
    assert item.display_config["section"] == "parameter"
    assert item.presentation_hash == validate_analysis_presentation_config(
        item.chart_config, item.display_config
    )

    with pytest.raises(DomainError) as tampered:
        repository._load_work_item(
            _StoredWorkItemConnection(tamper_presentation=True),  # type: ignore[arg-type]
            41,
            expected_lease_token="11111111-1111-4111-8111-111111111111",
        )
    assert tampered.value.code == "ANALYTICS_EXPORT_INTEGRITY_ERROR"


def test_sql_worker_rejects_inactive_requester_before_loading_export_payload(
    tmp_path: Path,
) -> None:
    repository = SqlAnalyticsExportWorkerRepository(
        object(),  # type: ignore[arg-type]
        AnalyticsExportPathPolicy(tmp_path / "exports"),
        worker_id="worker-test",
    )
    connection = _StoredWorkItemConnection(authorized=False)

    with pytest.raises(DomainError) as revoked:
        repository._load_work_item(
            connection,  # type: ignore[arg-type]
            41,
            expected_lease_token="11111111-1111-4111-8111-111111111111",
        )

    assert revoked.value.code == "ANALYTICS_EXPORT_ACCESS_REVOKED"
    authorization_sql = connection.calls[0]
    assert "iam.app_user access_user WITH (UPDLOCK,HOLDLOCK)" in authorization_sql
    assert "access_user.status='ACTIVE'" in authorization_sql
    assert "access_grant.status='ACTIVE'" in authorization_sql
    assert "access_domain.active=1" in authorization_sql
    assert "access_grant.expires_at_utc>SYSUTCDATETIME()" in authorization_sql
    assert "dataset.dataset_version dv WITH (UPDLOCK,HOLDLOCK)" in authorization_sql
    assert "dv.status='PUBLISHED'" in authorization_sql
    assert "dv.is_current=1" in authorization_sql
    assert not any(
        sql.startswith("SELECT export_job_id,requested_by,dataset_version_id")
        for sql in connection.calls
    )


def test_sql_worker_replays_typed_current_page_filters_and_legacy_v1(
    tmp_path: Path,
) -> None:
    repository = SqlAnalyticsExportWorkerRepository(
        object(),  # type: ignore[arg-type]
        AnalyticsExportPathPolicy(tmp_path / "exports"),
        worker_id="worker-test",
    )
    typed = repository._load_work_item(
        _StoredWorkItemConnection(  # type: ignore[arg-type]
            current_page=True,
            include_frozen_detail=True,
        ),
        41,
        expected_lease_token="11111111-1111-4111-8111-111111111111",
    )

    assert typed.current_page_detail_state is not None
    assert typed.current_page_detail_state.evaluation_filter is not None
    assert typed.current_page_detail_state.evaluation_filter.evaluation_type == "PAT"
    assert typed.current_page_detail_state.measurement_filter is not None
    assert typed.current_page_detail_state.measurement_filter.parameter == "IDSS"
    assert typed.current_page_detail_state.measurement_filter.upper_inclusive is False

    legacy = repository._load_work_item(
        _StoredWorkItemConnection(  # type: ignore[arg-type]
            current_page=True,
            include_frozen_detail=False,
        ),
        41,
        expected_lease_token="11111111-1111-4111-8111-111111111111",
    )
    assert legacy.current_page_detail_state == typed.current_page_detail_state

    with pytest.raises(DomainError) as tampered:
        repository._load_work_item(
            _StoredWorkItemConnection(  # type: ignore[arg-type]
                current_page=True,
                include_frozen_detail=True,
                tamper_frozen_detail=True,
            ),
            41,
            expected_lease_token="11111111-1111-4111-8111-111111111111",
        )
    assert tampered.value.code == "ANALYTICS_EXPORT_INTEGRITY_ERROR"


class _ClaimConnection:
    def __init__(self, state: _ClaimState) -> None:
        self.state = state

    def execute(self, statement, _parameters=None):
        sql = " ".join(str(statement).split())
        self.state.sql.append(sql)
        if sql.startswith(";WITH exhausted AS"):
            return _OptionalMappingResult()
        if sql.startswith(";WITH candidate AS"):
            if self.state.available:
                self.state.available = False
                return _OptionalMappingResult(
                    {
                        "export_job_id": 41,
                        "previous_status": "QUEUED",
                        "previous_lease_owner": None,
                        "previous_attempt_count": 0,
                        "attempt_count": 1,
                        "lease_expires_at_utc": datetime.now(UTC)
                        + timedelta(minutes=5),
                    }
                )
            return _OptionalMappingResult()
        if sql.startswith("SELECT COUNT_BIG(*) FROM delivery.export_artifact"):
            return _ScalarResult(0)
        if sql.startswith("INSERT governance.audit_log"):
            return _ScalarResult()
        raise AssertionError(sql)


class _ClaimContext(AbstractContextManager[_ClaimConnection]):
    def __init__(self, state: _ClaimState) -> None:
        self.connection = _ClaimConnection(state)

    def __enter__(self):
        return self.connection

    def __exit__(self, exc_type, exc_value, traceback):
        return None


class _ClaimEngine:
    def __init__(self, state: _ClaimState) -> None:
        self.state = state

    def begin(self):
        return _ClaimContext(self.state)


class _ClaimRepository(SqlAnalyticsExportWorkerRepository):
    def __init__(self, engine, policy, item) -> None:
        super().__init__(engine, policy, worker_id="worker-test")
        self.item = item

    def _load_work_item(self, _connection, _export_job_id, *, expected_lease_token):
        return replace(
            self.item,
            lease_token=expected_lease_token,
            lease_owner="worker-test",
            lease_expires_at_utc=datetime.now(UTC) + timedelta(minutes=5),
            attempt_count=1,
        )


def test_sql_claim_uses_skip_locked_semantics_and_only_one_worker_wins(
    tmp_path: Path,
) -> None:
    state = _ClaimState()
    engine = _ClaimEngine(state)
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    first = _ClaimRepository(engine, policy, _work_item()).claim_next()
    second = _ClaimRepository(engine, policy, _work_item()).claim_next()

    assert first is not None
    assert second is None
    claim_sql = next(sql for sql in state.sql if sql.startswith(";WITH candidate AS"))
    assert "ROWLOCK,READPAST,UPDLOCK" in claim_sql
    assert "status='QUEUED'" in claim_sql
    assert "status='RUNNING'" in claim_sql
    assert "lease_expires_at_utc<=SYSUTCDATETIME()" in claim_sql
    assert "attempt_count<max_attempts" in claim_sql
    assert "attempt_count=attempt_count+1" in claim_sql
    exhausted_sql = next(
        sql for sql in state.sql if sql.startswith(";WITH exhausted AS")
    )
    assert "attempt_count>=max_attempts" in exhausted_sql
    assert "ANALYTICS_EXPORT_RETRY_EXHAUSTED" in exhausted_sql


class _FailureConnection:
    def __init__(self) -> None:
        self.calls: list[tuple[str, dict[str, object]]] = []

    def execute(self, statement, parameters=None):
        sql = " ".join(str(statement).split())
        values = dict(parameters or {})
        self.calls.append((sql, values))
        if sql.startswith("UPDATE delivery.export_job SET status='FAILED'"):
            return _RowcountResult(1)
        if sql.startswith("INSERT governance.audit_log"):
            return _ScalarResult()
        raise AssertionError(sql)


class _FailureEngine:
    def __init__(self) -> None:
        self.connection = _FailureConnection()

    def begin(self):
        return nullcontext(self.connection)


def test_sql_repository_persists_stable_failed_reason(tmp_path: Path) -> None:
    engine = _FailureEngine()
    repository = SqlAnalyticsExportWorkerRepository(
        engine,  # type: ignore[arg-type]
        AnalyticsExportPathPolicy(tmp_path / "exports"),
        worker_id="worker-test",
    )

    repository.fail(
        _work_item(export_format="PDF"),
        error_code="ANALYTICS_EXPORT_DEPENDENCY_UNAVAILABLE",
        error_message="PDF rendering dependency is unavailable",
    )

    update_sql, parameters = engine.connection.calls[0]
    assert "finished_at_utc=SYSUTCDATETIME()" in update_sql
    assert parameters["error_message"] == (
        "ANALYTICS_EXPORT_DEPENDENCY_UNAVAILABLE: "
        "PDF rendering dependency is unavailable"
    )


class _MappingResult:
    def __init__(self, row: dict[str, object]) -> None:
        self.row = row

    def mappings(self):
        return self

    def one_or_none(self):
        return self.row


class _CompleteConnection:
    def __init__(self, *, authorized: bool = True) -> None:
        self.calls: list[tuple[str, dict[str, object]]] = []
        self.authorized = authorized

    def execute(self, statement, parameters=None):
        sql = " ".join(str(statement).split())
        values = dict(parameters or {})
        self.calls.append((sql, values))
        if sql.startswith("SELECT status,lease_token"):
            return _MappingResult(
                {
                    "status": "RUNNING",
                    "lease_token": "11111111-1111-4111-8111-111111111111",
                    "lease_owner": "worker-test",
                    "lease_expires_at_utc": datetime.now(UTC) + timedelta(minutes=5),
                }
            )
        if sql.startswith("SELECT ejd.dataset_version_id"):
            return _MappingRowsResult(
                [
                    {
                        "dataset_version_id": 71,
                        "ordinal_no": 1,
                        "dataset_id": 7,
                        "version_no": 3,
                        "status": "PUBLISHED",
                        "is_current": True,
                        "test_stage": "CP",
                        "requested_by_user_id": 10,
                        "job_status": "RUNNING",
                        "job_contract_version": "ANALYTICS_EXPORT_V1",
                        "job_lease_token": (
                            "11111111-1111-4111-8111-111111111111"
                        ),
                        "job_lease_owner": "worker-test",
                        "job_lease_expires_at_utc": datetime.now(UTC)
                        + timedelta(minutes=5),
                        "can_read": self.authorized,
                    }
                ]
            )
        if sql.startswith("SELECT COUNT_BIG(*) FROM delivery.export_artifact"):
            return _ScalarResult(0)
        if sql.startswith("INSERT delivery.export_artifact"):
            return _ScalarResult()
        if sql.startswith("UPDATE delivery.export_job SET status='SUCCESS'"):
            return _RowcountResult(1)
        if sql.startswith("INSERT governance.audit_log"):
            return _ScalarResult()
        raise AssertionError(sql)


class _CompleteEngine:
    def __init__(self) -> None:
        self.connection = _CompleteConnection()

    def begin(self):
        return nullcontext(self.connection)


class _HeartbeatConnection:
    def __init__(self, rowcount: int) -> None:
        self.rowcount = rowcount
        self.calls: list[tuple[str, dict[str, object]]] = []

    def execute(self, statement, parameters=None):
        sql = " ".join(str(statement).split())
        values = dict(parameters or {})
        self.calls.append((sql, values))
        return _RowcountResult(self.rowcount)


class _HeartbeatEngine:
    def __init__(self, rowcount: int) -> None:
        self.connection = _HeartbeatConnection(rowcount)

    def begin(self):
        return nullcontext(self.connection)


def test_repository_heartbeat_is_token_owner_and_expiry_fenced(
    tmp_path: Path,
) -> None:
    engine = _HeartbeatEngine(1)
    repository = SqlAnalyticsExportWorkerRepository(
        engine,  # type: ignore[arg-type]
        AnalyticsExportPathPolicy(tmp_path / "exports"),
        worker_id="worker-test",
        lease_seconds=120,
    )
    item = _work_item()

    repository.heartbeat(item)

    sql, parameters = engine.connection.calls[0]
    assert "lease_token=:lease_token" in sql
    assert "lease_owner=:lease_owner" in sql
    assert "lease_expires_at_utc>=SYSUTCDATETIME()" in sql
    assert parameters["lease_token"] == item.lease_token
    assert parameters["lease_seconds"] == 120

    lost = SqlAnalyticsExportWorkerRepository(
        _HeartbeatEngine(0),  # type: ignore[arg-type]
        AnalyticsExportPathPolicy(tmp_path / "lost"),
        worker_id="worker-test",
    )
    with pytest.raises(DomainError) as failure:
        lost.heartbeat(item)
    assert failure.value.code == "ANALYTICS_EXPORT_WORKER_CLAIM_LOST"


class _StaleCompleteConnection(_CompleteConnection):
    def execute(self, statement, parameters=None):
        sql = " ".join(str(statement).split())
        values = dict(parameters or {})
        self.calls.append((sql, values))
        if sql.startswith("SELECT status,lease_token"):
            return _MappingResult(
                {
                    "status": "RUNNING",
                    "lease_token": "22222222-2222-4222-8222-222222222222",
                    "lease_owner": "replacement-worker",
                    "lease_expires_at_utc": datetime.now(UTC) + timedelta(minutes=5),
                }
            )
        raise AssertionError(sql)


class _StaleCompleteEngine:
    def __init__(self) -> None:
        self.connection = _StaleCompleteConnection()

    def begin(self):
        return nullcontext(self.connection)


def test_old_worker_cannot_register_artifact_after_stale_recovery(
    tmp_path: Path,
) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    item = _work_item()
    artifact = AnalyticsExportRenderer(policy, _Content()).render(item)
    engine = _StaleCompleteEngine()

    with pytest.raises(DomainError) as failure:
        SqlAnalyticsExportWorkerRepository(
            engine,  # type: ignore[arg-type]
            policy,
            worker_id="worker-test",
        ).complete(
            item,
            artifact,
            expires_at_utc=datetime.now(UTC) + timedelta(hours=1),
        )

    assert failure.value.code == "ANALYTICS_EXPORT_WORKER_CLAIM_LOST"
    assert not any(
        sql.startswith("INSERT delivery.export_artifact")
        for sql, _ in engine.connection.calls
    )


def test_revocation_during_render_blocks_success_artifact_registration(
    tmp_path: Path,
) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    item = _work_item()
    artifact = AnalyticsExportRenderer(policy, _Content()).render(item)
    engine = _CompleteEngine()
    engine.connection.authorized = False

    with pytest.raises(DomainError) as revoked:
        SqlAnalyticsExportWorkerRepository(
            engine,  # type: ignore[arg-type]
            policy,
            worker_id="worker-test",
        ).complete(
            item,
            artifact,
            expires_at_utc=datetime.now(UTC) + timedelta(hours=1),
        )

    assert revoked.value.code == "ANALYTICS_EXPORT_ACCESS_REVOKED"
    assert not any(
        sql.startswith("INSERT delivery.export_artifact")
        for sql, _ in engine.connection.calls
    )
    assert not any(
        sql.startswith("UPDATE delivery.export_job SET status='SUCCESS'")
        for sql, _ in engine.connection.calls
    )
    authorization_sql = next(
        sql
        for sql, _ in engine.connection.calls
        if sql.startswith("SELECT ejd.dataset_version_id")
    )
    assert "iam.app_user access_user WITH (UPDLOCK,HOLDLOCK)" in authorization_sql
    assert "access_user.status='ACTIVE'" in authorization_sql
    assert "dataset.dataset_version dv WITH (UPDLOCK,HOLDLOCK)" in authorization_sql
    assert "dv.status='PUBLISHED'" in authorization_sql
    assert "dv.is_current=1" in authorization_sql


def test_requester_disabled_after_render_blocks_success_registration(
    tmp_path: Path,
) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    item = _work_item(export_job_id=142)
    artifact = AnalyticsExportRenderer(policy, _Content()).render(item)
    engine = _CompleteEngine()
    engine.connection.authorized = False

    with pytest.raises(DomainError) as disabled:
        SqlAnalyticsExportWorkerRepository(
            engine,  # type: ignore[arg-type]
            policy,
            worker_id="worker-test",
        ).complete(
            item,
            artifact,
            expires_at_utc=datetime.now(UTC) + timedelta(hours=1),
        )

    assert disabled.value.code == "ANALYTICS_EXPORT_ACCESS_REVOKED"
    assert not any(
        sql.startswith("INSERT delivery.export_artifact")
        for sql, _ in engine.connection.calls
    )


@pytest.mark.parametrize(
    ("export_format", "mime_type"),
    (("PDF", "application/pdf"), ("PNG", "image/png")),
)
def test_repository_transactionally_registers_real_report_media(
    tmp_path: Path, export_format: str, mime_type: str
) -> None:
    item = _work_item(
        export_format=export_format,
        export_scope="REPORT",
        template_code="ANALYTICS_OVERVIEW",
        export_job_id=52,
    )
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    artifact = AnalyticsExportRenderer(policy, _Content()).render(item)
    engine = _CompleteEngine()

    SqlAnalyticsExportWorkerRepository(
        engine,  # type: ignore[arg-type]
        policy,
        worker_id="worker-test",
    ).complete(
        item,
        artifact,
        expires_at_utc=datetime.now(UTC) + timedelta(hours=1),
    )

    artifact_insert = next(
        values
        for sql, values in engine.connection.calls
        if sql.startswith("INSERT delivery.export_artifact")
    )
    assert artifact_insert["mime_type"] == mime_type
    assert artifact_insert["file_size"] == artifact.file_size
    assert artifact_insert["sha256"] == artifact.sha256
    assert artifact_insert["storage_uri"] == str(artifact.path)
    finalize_sql = next(
        sql
        for sql, _values in engine.connection.calls
        if sql.startswith("UPDATE delivery.export_job SET status='SUCCESS'")
    )
    assert "iam.app_user access_user WITH (UPDLOCK,HOLDLOCK)" in finalize_sql
    assert "access_user.status='ACTIVE'" in finalize_sql
    assert "finalize_dv.status='PUBLISHED'" in finalize_sql
    assert "finalize_dv.is_current=1" in finalize_sql
    assert "finalize_dv WITH (UPDLOCK,HOLDLOCK)" in finalize_sql


def test_repository_rejects_expired_or_tampered_artifact_before_sql(
    tmp_path: Path,
) -> None:
    policy = AnalyticsExportPathPolicy(tmp_path / "exports")
    item = _work_item()
    policy.prepare_job_root(item.export_job_id)
    path = policy.artifact_path(item.export_job_id, "analytics-export-41.csv")
    path.write_bytes(b"content")
    identity = policy.identify(item.export_job_id, path)
    artifact = RenderedAnalyticsExport(
        path=path,
        file_name=path.name,
        mime_type="text/csv; charset=utf-8",
        file_size=identity.file_size,
        sha256=identity.sha256,
        exported_row_count=1,
    )
    repository = SqlAnalyticsExportWorkerRepository(
        _ClaimEngine(_ClaimState()), policy, worker_id="worker-test"
    )

    with pytest.raises(DomainError) as expired:
        repository.complete(
            item,
            artifact,
            expires_at_utc=datetime.now(UTC) - timedelta(seconds=1),
        )
    assert expired.value.code == "ANALYTICS_EXPORT_TTL_INVALID"

    with pytest.raises(DomainError) as tampered:
        repository.complete(
            item,
            replace(artifact, sha256="0" * 64),
            expires_at_utc=datetime.now(UTC) + timedelta(hours=1),
        )
    assert tampered.value.code == "ANALYTICS_EXPORT_ARTIFACT_INTEGRITY_ERROR"


def test_download_metadata_blocks_expired_artifact() -> None:
    item = _work_item()
    now = datetime.now(UTC)
    record = AnalyticsExportRecord(
        export_job_id=item.export_job_id,
        requested_by=item.requested_by,
        contract_version="ANALYTICS_EXPORT_V1",
        worker_contract_version="ANALYTICS_EXPORT_WORKER_V1",
        generation_mode="QUEUED_WORKER",
        status="SUCCESS",
        export_scope=item.export_scope.value,
        export_format=item.export_format.value,
        template_code=item.template_code,
        template_version=item.template_version,
        datasets=(AnalyticsExportDatasetRecord(71, 7, 3, 1, "CP"),),
        filters={},
        parameters=("IDSS",),
        filter_hash=item.filter_hash,
        context_hash=item.context_hash,
        rule_context=item.rule_context,
        chart_config=item.chart_config,
        display_config=item.display_config,
        presentation_hash=item.presentation_hash,
        artifact_ttl_hours=24,
        page=None,
        page_size=None,
        idempotency_key="analytics-export-test",
        request_reason_sha256="c" * 64,
        requested_at_utc=now.isoformat(),
        started_at_utc=now.isoformat(),
        finished_at_utc=now.isoformat(),
        exported_row_count=1,
        row_version="0000000000000001",
        idempotent_replay=False,
    )
    result = SqlAnalyticsExportService._download_result(
        record,
        (
            {
                "export_artifact_id": 1,
                "file_name": "analytics-export-41.csv",
                "mime_type": "text/csv; charset=utf-8",
                "file_size": 10,
                "sha256": "d" * 64,
                "created_at_utc": now - timedelta(hours=2),
                "expires_at_utc": now - timedelta(hours=1),
            },
        ),
    )

    assert result.availability == "EXPIRED"
    assert result.download_enabled is False
    assert result.reason_code == "ANALYTICS_EXPORT_ARTIFACT_EXPIRED"

    valid_row = {
        "export_artifact_id": 2,
        "file_name": "analytics-export-41.csv",
        "mime_type": "text/csv; charset=utf-8",
        "file_size": 10,
        "sha256": "d" * 64,
        "created_at_utc": now,
        "expires_at_utc": now + timedelta(hours=1),
    }
    duplicate = SqlAnalyticsExportService._download_result(
        record, (valid_row, {**valid_row, "export_artifact_id": 3})
    )
    assert duplicate.availability == "INTEGRITY_BLOCKED"
    assert duplicate.download_enabled is False

    cleaned = SqlAnalyticsExportService._download_result(
        replace(record, status="EXPIRED"),
        ({**valid_row, "physical_status": "DELETED"},),
    )
    assert cleaned.availability == "EXPIRED"
    assert cleaned.download_enabled is False
    assert cleaned.reason_code == "ANALYTICS_EXPORT_EXPIRED"
