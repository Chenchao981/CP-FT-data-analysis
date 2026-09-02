from __future__ import annotations

import hashlib
import subprocess
from pathlib import Path

import pytest
from app.domain.cleaner_registry import CleanerRelease
from app.infrastructure.existing_cleaner_runner import (
    CleanerArtifact,
    ExistingCleanerRunResult,
)
from app.infrastructure.quick_pat_runner import QuickPatRunner
from app.infrastructure.source_catalog import SourceCatalog, SourceRoot
from openpyxl import Workbook


def _release(package: Path, runtime: Path) -> CleanerRelease:
    return CleanerRelease(
        21,
        8,
        "FT",
        "JIEQUN",
        "JIEQUN_FT_QUICK_PAT_EXISTING",
        "route-a-v1",
        "JIEQUN_FT_QUICK_PAT_EXISTING",
        "sha256-test",
        hashlib.sha256(package.read_bytes()).hexdigest(),
        str(package),
        str(runtime),
        "factories.jiequn.pat_cleaner.generate_raw_pat",
        "JIEQUN_FT_QUICK_PAT_PYZ",
        "JIEQUN_UNIFIED_CSV_DIRECTORY_V1",
        "FT_PAT_RESULT_V1",
        None,
        30,
        10_000_000,
    )


def test_runner_invokes_released_package_and_builds_auditable_artifacts(
    monkeypatch, tmp_path: Path,
) -> None:
    package = tmp_path / "ft_data_cleaner.pyz"
    runtime = tmp_path / "python.exe"
    package.write_bytes(b"released-pat")
    runtime.touch()
    source = tmp_path / "source" / "product"
    source.mkdir(parents=True)
    (source / "one.csv").write_text("x\n1\n", encoding="utf-8")
    catalog = SourceCatalog(
        (
            SourceRoot(
                "ROOT", "Root", tmp_path / "source", "FT", "JIEQUN", (".csv",)
            ),
        )
    )
    manifest = catalog.build_manifest("ROOT", "product")
    output = tmp_path / "output"
    monkeypatch.setenv("TMS_DATABASE_URL", "must-not-reach-cleaner")
    monkeypatch.setenv("TMS_JWT_SECRET", "must-not-reach-cleaner")

    def fake_run(command, **kwargs):
        assert command[:2] == [str(runtime), "-c"]
        assert "generate_raw_pat" in command[2]
        assert kwargs["env"]["TMS_FT_PAT_INPUT"] == str(source.resolve())
        assert kwargs["env"]["TMS_FT_PAT_ADAPTER"] == "JIEQUN_FT_QUICK_PAT_PYZ"
        assert "TMS_DATABASE_URL" not in kwargs["env"]
        assert "TMS_JWT_SECRET" not in kwargs["env"]
        report = output / "PAT_001" / "PAT_001.xlsx"
        report.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["统计量", "总计数", "均值", "标准差"])
        sheet.append(["变量", "总计数", "均值", "标准差"])
        sheet.append(["VTH", 100, 4.1, 0.2])
        sheet.append(["RDON", 98, 0.01, 0.001])
        workbook.save(report)
        return subprocess.CompletedProcess(
            command,
            0,
            stdout="杰群 PAT 原始数据汇总完成: 文件=1, 解析行=100, 参数=2\n",
            stderr="",
        )

    result = QuickPatRunner(process_runner=fake_run).run_release(
        release=_release(package, runtime),
        input_directory=source,
        output_root=output,
        source_manifest_json=manifest.as_json(),
        source_manifest_sha256=manifest.sha256,
    )
    assert result.parameter_count == 2
    assert result.record_count == 100
    assert {item.role for item in result.artifacts} == {
        "pat_report",
        "pat_summary",
        "source_manifest",
    }
    assert result.summary["parameters"][1]["parameter"] == "RDON"
    assert all(len(item.sha256) == 64 for item in result.artifacts)


@pytest.mark.parametrize(
    ("factory", "adapter", "input_contract", "source_name"),
    [
        (
            "RIYUEXIN",
            "RIYUEXIN_FT_QUICK_PAT_PYZ",
            "RIYUEXIN_RAW_XLSX_DIRECTORY_V1",
            "wafer.xlsx",
        ),
        (
            "DIANJI",
            "DIANJI_FT_QUICK_PAT_PYZ",
            "DIANJI_REGISTERED_RAW_DIRECTORY_V1",
            "raw.xls",
        ),
    ],
)
def test_runner_dispatches_additional_released_ft_pat_adapters(
    tmp_path: Path,
    factory: str,
    adapter: str,
    input_contract: str,
    source_name: str,
) -> None:
    package = tmp_path / "ft_data_cleaner.pyz"
    runtime = tmp_path / "python.exe"
    package.write_bytes(b"released-ft")
    runtime.touch()
    source = tmp_path / "source"
    source.mkdir()
    (source / source_name).write_bytes(b"raw")
    manifest_payload = (
        '{"file_count":1,"files":[],"mode":"LOCAL_PATH_SIZE_MTIME_V1",'
        '"source_label":"source","total_bytes":3}'
    )
    manifest_sha = hashlib.sha256(manifest_payload.encode("utf-8")).hexdigest()
    output = tmp_path / "output"
    release = CleanerRelease(
        41, 10, "FT", factory, f"{factory}_FT_QUICK_PAT_EXISTING", "route-a-v1",
        f"{factory}_FT_QUICK_PAT_EXISTING", "v1",
        hashlib.sha256(package.read_bytes()).hexdigest(), str(package), str(runtime),
        "generate_raw_pat", adapter, input_contract, "FT_PAT_RESULT_V1", None,
        30, 10_000_000,
    )

    def fake_run(command, **kwargs):
        assert command[:2] == [str(runtime), "-c"]
        assert "generate_raw_pat" in command[2]
        assert kwargs["env"]["TMS_FT_PAT_ADAPTER"] == adapter
        report = output / "PAT_001" / "PAT_001.xlsx"
        report.parent.mkdir(parents=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["statistics", "count"])
        sheet.append(["variable", "count"])
        sheet.append(["VTH", 100])
        workbook.save(report)
        return subprocess.CompletedProcess(
            command,
            0,
            stdout=f"{factory} PAT raw data summary: 文件=1, 解析行=100, 参数=1\n",
            stderr="",
        )

    result = QuickPatRunner(process_runner=fake_run).run_release(
        release=release,
        input_directory=source,
        output_root=output,
        source_manifest_json=manifest_payload,
        source_manifest_sha256=manifest_sha,
    )

    assert result.summary["factory_code"] == factory
    assert result.summary["formula_contract"] == (
        "SIGMA_IQR_1_35_MEDIAN_PLUS_MINUS_6SIGMA_V1"
    )
    assert result.record_count == 100


def test_runner_uses_cp_release_for_cleaning_and_package_owned_pat(tmp_path: Path) -> None:
    package = tmp_path / "app.pyz"
    runtime = tmp_path / "python.exe"
    package.write_bytes(b"released-cp")
    runtime.touch()
    source = tmp_path / "source"
    source.mkdir()
    (source / "wafer.xlsx").write_bytes(b"source")
    manifest_payload = (
        '{"file_count":1,"files":[],"mode":"LOCAL_PATH_SIZE_MTIME_V1",'
        '"source_label":"source","total_bytes":6}'
    )
    manifest_sha = hashlib.sha256(manifest_payload.encode("utf-8")).hexdigest()
    output = tmp_path / "output"
    release = CleanerRelease(
        31, 9, "CP", "JETECH", "JETECH_CP_EXISTING", "route-a-v1",
        "JETECH_CP_EXISTING", "v1", hashlib.sha256(package.read_bytes()).hexdigest(),
        str(package), str(runtime), "cleaner + quick_pat", "JETECH_CP_PYZ",
        "CP_EXCEL_OR_ZIP_V1", "CP_STANDARD_CSV_TRIPLET_V1", None, 30, 10_000_000,
    )

    class StubCleanerRunner:
        def run_release(self, **kwargs):
            assert kwargs["release"] == release
            assert kwargs["inputs"] == (source.resolve(),)
            intermediate = Path(kwargs["output_root"])
            intermediate.mkdir(parents=True)
            cleaned = intermediate / "LOT_cleaned_1.csv"
            spec = intermediate / "LOT_spec_1.csv"
            cleaned.write_text("Lot_ID,VTH\nLOT,1\n", encoding="utf-8")
            spec.write_text("Parameter,Unit\nVTH,V\n", encoding="utf-8")
            artifacts = (
                CleanerArtifact("cleaned", str(cleaned), cleaned.stat().st_size, "a" * 64),
                CleanerArtifact("spec", str(spec), spec.stat().st_size, "b" * 64),
            )
            return ExistingCleanerRunResult("CP", "jetech", str(intermediate), artifacts, "cleaned")

    def fake_run(command, **kwargs):
        assert command[:2] == [str(runtime), "-c"]
        assert "generate_cleaned_csv_pat" in command[2]
        report = output / "PAT_CP_20260902_120000.xlsx"
        report.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["CP PAT", "contract"])
        sheet.append(["parameter", "count", "mean", "stddev"])
        sheet.append(["VTH", 10, 4.1, 0.2])
        workbook.save(report)
        stdout = (
            'TMS_CP_PAT_SUMMARY={"source_files":1,"record_count":10,'
            '"parameter_count":1,"formula_contract":"AEC_Q101_MEDIAN_IQR_5SIGMA_VDMOS_V5_6",'
            f'"report":"{report.as_posix()}"}}\n'
        )
        return subprocess.CompletedProcess(command, 0, stdout=stdout, stderr="")

    result = QuickPatRunner(
        process_runner=fake_run, cleaner_runner=StubCleanerRunner()
    ).run_release(
        release=release,
        input_directory=source,
        output_root=output,
        source_manifest_json=manifest_payload,
        source_manifest_sha256=manifest_sha,
    )
    assert result.parameter_count == 1
    assert result.record_count == 10
    assert result.summary["test_stage"] == "CP"
    assert result.summary["factory_code"] == "JETECH"
    assert not (output / "cp_cleaner_intermediate").exists()
